# -*- coding: utf-8 -*-
"""
公共工程施工日誌產生器（通用版）

用途：
  給任何一份 PCCES 匯出的「詳細價目表 / 標價清單 / 預算詳細表」（.xls 或 .xlsx），
  自動產生一份完整的施工管理活頁簿：
    日資料庫（80/N個工項逐日填報表）／預定進度（任務排程→逐日累計%）／
    甘特圖（橫向長條+S曲線）／施工日報（依公共工程委員會附表四官方格式）
  不會把任何工項數量/大項中項邊界寫死，全部從您提供的
  PCCES 檔案內容自動偵測，所以理論上換一個新案子直接餵新檔案就能用。

安裝（只需要做一次）：
  1. 安裝 Python 3.9 以上版本（https://www.python.org/downloads/，安裝時記得
     勾選「Add python.exe to PATH」）。
  2. 打開命令提示字元(cmd)，切到這個資料夾，執行：
       pip install -r requirements.txt

使用方式：
  python build_construction_log.py "您的PCCES檔案路徑.xls"

  常用參數（都有預設值，不給也能跑）：
    --sheet 工作表名稱        手動指定要讀哪個分頁（預設自動找含「項次」表頭的分頁）
    --start-date 2026-09-01   開工日期（預設今天，事後也可以直接在Excel B2改）
    --days 200                日資料庫要建幾天份（預設200天）
    --agency "○○縣政府"       主辦機關（PCCES檔案讀不到時可用這個手動補）
    --out 輸出檔名.xlsx        輸出檔名（預設用工程名稱自動命名）
    --n-mat / --n-labor / --n-mach   材料/工別/機具主檔筆數上限（預設12/8/8）

  範例：
    python build_construction_log.py "案A_標價清單.xlsx" --start-date 2026-09-01 --out 案A_施工日誌.xlsx

輸出：
  同資料夾下產生 <out>.xlsx（4張核心分頁＋1張「PCCES來源核對(debug用)」分頁）與
  <out>_meta.json（結構資訊，之後如果要用「估驗計價工具」「變更設計匯入工具」都需要
  讀這個meta檔）。

  「PCCES來源核對(debug用)」分頁：把程式實際解析出來的每一列原始資料（含大項/中項/
  小計/總計標題列，不只是N個工項本身）、契約總價是用哪一列文字判斷出來的、以及大項/
  中項偵測結果，整理成一張表，方便您拿這張表跟自己手上的PCCES畫面逐列核對，快速看
  出是不是選錯分頁、漏抓欄位、或編號沒偵測到——不是官方施工日誌的固定分頁，不需要
  就可以刪掉或隱藏。

已知限制（跟真人核對過的假設，若您的案子不符合請留意）：
  - 假設 PCCES 表格是標準7欄格式：項次/項目及說明/單位/數量/單價/複價/編碼(備註)，
    順序不能變動（這是 PCCES 軟體標準匯出格式，一般不會不同）。
  - 大項用「壹貳參肆伍…」、中項用「一二三四五…」這兩組中文數字番號辨識階層，
    如果您的PCCES用別種編號方式（例如阿拉伯數字1.2.3.或英文字母），偵測不到時
    大項/中項清單會是空的，仍會照樣把80個(或N個)工項匯入，只是沒有分類進度欄，
    不影響逐日填報功能本身，只是「施工進度統計」表格會顯示不出分類小計。
  - 每個大項/中項的「契約金額」一律用「該範圍內所有工項的複價加總」自行核算，
    不直接信任PCCES檔案裡的「小計」「總價」文字列本身寫的數字。
  - 契約總價的判斷順序：優先找文字裡「有+號」的總價列（例如「總價(壹+貳+參+肆+伍)」，
    這種寫法最明確、視為最終總價）；找不到的話，改用文件裡「最後一筆」總價列；兩種
    都找不到才退回全部工項複價自行加總。不論用哪一種，都會另外把「全部工項複價自行
    加總」的數字拿來跟採用值互相核對，兩者對不起來就印警告並在「PCCES來源核對」分頁
    列出差額——常見於契約文件在「總價」列之後還接了一個另計的扣款/調整項次（例如
    「剩餘價值折價費」，依業務慣例不算入契約金額），這種情況請以「PCCES來源核對」
    分頁列出的差額原因為準，必要時用 --project-name 等參數或直接編輯產生的Excel手動
    修正。
"""
import argparse
import datetime
import json
import os
import re
import sys

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.chart import LineChart, Reference
from openpyxl.worksheet.properties import Outline
from openpyxl.worksheet.table import Table, TableStyleInfo

TOP_NUMERALS = ['壹', '貳', '參', '肆', '伍', '陸', '柒', '捌', '玖', '拾',
                '拾壹', '拾貳', '拾參', '拾肆', '拾伍', '拾陸', '拾柒', '拾捌', '拾玖', '貳拾']
MID_NUMERALS = ['一', '二', '三', '四', '五', '六', '七', '八', '九', '十',
                 '十一', '十二', '十三', '十四', '十五', '十六', '十七', '十八', '十九', '二十',
                 '廿一', '廿二', '廿三', '廿四', '廿五']

FONT_NAME = '新細明體'

# 人員(工別)/機具主檔的預設名稱清單：使用者要求的常用工班/機具名稱，先幫忙預填在
# 主檔的前幾格（超過清單長度的欄位仍留空白給使用者自行增填）。單位欄採用公共工程
# 慣例：人員以「工」計、機具以「台班」計——這是常見慣例但非強制規定，若跟貴案的
# 契約單位寫法不同，請直接在Excel裡覆寫這幾格即可，不影響公式（單位欄只是顯示用，
# 沒有任何公式引用它）。
DEFAULT_LABOR_NAMES = ['工程師', '一般技工', '生產體力工', '普通工', '技術工']
DEFAULT_LABOR_UNIT = '工'
DEFAULT_MACH_NAMES = ['挖土機', '傾卸貨車', '吊(卡)車', '鑽掘機']
DEFAULT_MACH_UNIT = '台班'

# 日資料庫「八大項文字欄+預定進度%」的固定欄位定義。搬到模組層級（而非只在
# build_day_database()裡當區域變數）是因為套用契約變更(apply_change_order)需要在完全
# 不知道new ctx的情況下，先用同一份固定的「完整說明文字」去舊檔案裡逐欄比對定位——
# 兩處共用同一份常數，才能保證版本不會日後改一邊忘了改另一邊、比對邏輯永遠對得起來。
TXT_FIELDS = [
    ('SEC1_A', '一、營造業專業工程特定施工項目 A.', 'text'),
    ('SEC1_B', '一、營造業專業工程特定施工項目 B.', 'text'),
    ('SEC4', '四、本日施工項目是否有須依「營造業專業工程特定施工項目應置之技術士種類、比率或人數標準表」規定應設置技術士之專業工程', 'yn'),
    ('SEC5A', '五-(一)-1 實施勤前教育（含工地預防災變及危害告知）', 'yn'),
    ('SEC5B', '五-(一)-2 確認新進勞工是否提報勞工保險（或其他商業保險）資料及安全衛生教育訓練紀錄', 'yn2'),
    ('SEC5C', '五-(一)-3 檢查勞工個人防護具', 'yn'),
    ('SEC5_OTHER', '五-(二) 其他事項', 'text'),
    ('SEC6', '六、施工取樣試驗紀錄：', 'text'),
    ('SEC7', '七、通知協力廠商辦理事項：', 'text'),
    ('SEC8', '八、重要事項記錄：', 'text'),
    ('SCHED', '預定進度(%)（自動：抓「預定進度」分頁依任務起訖日計算之逐日累計%，不需人工填寫）', 'num'),
]


# ============================================================
# 第一階段：讀取 PCCES 檔案 -> 一列一項的工項清單 + 大項/中項階層
# ============================================================

def _is_blank(v):
    return v is None or (isinstance(v, str) and v.strip() == '')


def _fmt(v):
    return '' if _is_blank(v) else v


def _read_rows_xls(path, sheet_name=None):
    import xlrd
    wb = xlrd.open_workbook(path)
    sh = wb.sheet_by_name(sheet_name) if sheet_name else _pick_sheet_xls(wb)
    rows = [[sh.cell_value(r, c) for c in range(sh.ncols)] for r in range(sh.nrows)]
    return rows, sh.name


def _is_detail_header_row(row):
    """判斷某一列是否為PCCES「詳細價目表」的表頭列（而不是預算總表/單價分析表等
    也同樣在A欄寫著「項次」的其他分頁）。真正的詳細表表頭在單位/數量/單價三欄
    都會有對應標籤；只用A欄「項次」判斷會誤選到預算總表（其欄位其實是金額彙總，
    B~E欄是空的）。"""
    def norm(v):
        return str(v).replace(' ', '').replace('　', '').strip() if isinstance(v, str) else ''
    if norm(row[0] if len(row) > 0 else None) not in ('項次',):
        return False
    unit = norm(row[2] if len(row) > 2 else None)
    qty = norm(row[3] if len(row) > 3 else None)
    price = norm(row[4] if len(row) > 4 else None)
    return unit == '單位' and qty == '數量' and price == '單價'


def _pick_sheet_xls(wb):
    fallback = None
    for name in wb.sheet_names():
        sh = wb.sheet_by_name(name)
        for r in range(min(15, sh.nrows)):
            row = [sh.cell_value(r, c) for c in range(sh.ncols)]
            if _is_detail_header_row(row):
                return sh
            v = row[0] if row else None
            if fallback is None and isinstance(v, str) and v.strip() in ('項次', '項 次'):
                fallback = sh
    return fallback if fallback is not None else wb.sheet_by_index(0)


def _read_rows_xlsx(path, sheet_name=None):
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = _pick_sheet_xlsx(wb)
    rows = [[c.value for c in row] for row in ws.iter_rows()]
    return rows, ws.title


def _pick_sheet_xlsx(wb):
    fallback = None
    for ws in wb.worksheets:
        for r in range(1, min(15, ws.max_row) + 1):
            row = [ws.cell(row=r, column=c).value for c in range(1, min(ws.max_column, 7) + 1)]
            if _is_detail_header_row(row):
                return ws
            v = row[0] if row else None
            if fallback is None and isinstance(v, str) and v.strip() in ('項次', '項 次'):
                fallback = ws
    return fallback if fallback is not None else wb.worksheets[0]


def read_pcces(path, sheet_name=None):
    """回傳 (rows, 實際使用的分頁名稱)，rows 是 0-index 的 list of list。"""
    ext = os.path.splitext(path)[1].lower()
    if ext == '.xls':
        return _read_rows_xls(path, sheet_name)
    elif ext in ('.xlsx', '.xlsm'):
        return _read_rows_xlsx(path, sheet_name)
    else:
        raise ValueError(f'不支援的副檔名：{ext}（只支援 .xls / .xlsx / .xlsm）')


def find_header_row(rows, max_scan=15):
    for r in range(min(max_scan, len(rows))):
        v = rows[r][0] if rows[r] else None
        if isinstance(v, str) and v.strip() in ('項次', '項 次'):
            return r
    raise ValueError('在前15列都找不到「項次」表頭，請確認選對分頁（用 --sheet 指定），'
                      '或這份檔案不是標準PCCES詳細價目表格式。')


def extract_header_meta(rows, header_row):
    """在表頭列之前掃描「工程名稱/工程地點/工程編號/主辦機關」等標籤，
    抓標籤右邊(或往右找到的第一個非空)儲存格當值。抓不到就回傳 None，
    由呼叫端決定要用固定位置備援或請使用者用CLI參數手動補。"""
    meta = {'project_name': None, 'location': None, 'proj_code': None, 'agency': None}
    label_map = {
        '工程名稱': 'project_name', '工程地點': 'location', '施工地點': 'location',
        '工程編號': 'proj_code', '主辦機關': 'agency', '機關名稱': 'agency',
    }
    for r in range(min(header_row, len(rows))):
        row = rows[r]
        for c, cell in enumerate(row):
            if not isinstance(cell, str):
                continue
            key = label_map.get(cell.strip())
            if not key or meta[key] is not None:
                continue
            for cc in range(c + 1, min(c + 6, len(row))):
                v = row[cc]
                if not _is_blank(v):
                    meta[key] = v
                    break
    # 備援：agency 常常出現在檔案最左上角(A1儲存格本身就是機關全銜)
    if meta['agency'] is None and rows and isinstance(rows[0][0], str) and not _is_blank(rows[0][0]):
        meta['agency'] = rows[0][0]
    return meta


def parse_items(rows, header_row):
    """把換列的工項合併成一列一項，回傳 records（含大項/中項標題列、小計/總計列、工項列）。"""
    records, cur = [], None
    ncols = max(len(r) for r in rows) if rows else 7
    ncols = max(ncols, 7)
    for r in range(header_row + 1, len(rows)):
        row = rows[r] + [None] * (ncols - len(rows[r]))
        a, b, c, d, e, f, g = row[0], row[1], row[2], row[3], row[4], row[5], row[6]
        if all(_is_blank(x) for x in (a, b, c, d, e, f, g)):
            continue
        starts_new = (not _is_blank(a)) or (_is_blank(a) and not _is_blank(f))
        if starts_new:
            if cur:
                records.append(cur)
            cur = {'項次': _fmt(a), '項目及說明': _fmt(b), '單位': _fmt(c),
                   '數量': _fmt(d), '單價': _fmt(e), '複價': _fmt(f), '編碼備註': _fmt(g)}
        else:
            if cur is None:
                continue
            if not _is_blank(b):
                cur['項目及說明'] = str(cur['項目及說明']) + str(b)
            if not _is_blank(g):
                cur['編碼備註'] = str(cur['編碼備註']) + str(g)
    if cur:
        records.append(cur)

    def classify(rec):
        text = str(rec['項目及說明'])
        has_qty = (not _is_blank(rec['單位'])) and (not _is_blank(rec['數量'])) and (not _is_blank(rec['單價']))
        if text.startswith('總價'):
            return '總計'
        if text.startswith('小計'):
            return '小計'
        if has_qty:
            return '工項'
        return '大項/中項'

    for rec in records:
        rec['類型'] = classify(rec)
    return records


def _item_amount(it):
    return it['複價'] if isinstance(it['複價'], (int, float)) else 0.0


def build_categories(records, items, N_ITEMS, grand_total):
    """通用大項/中項偵測：用「壹貳參…」「一二三…」這兩組番號在原始資料裡的出現順序
    直接切出彼此不重疊的區間，不假設固定的工項數邊界。每個類別的契約金額一律用
    「該區間內所有工項複價自行加總」得出，不信任PCCES檔案裡「小計/總價」列本身
    寫的數字（那個數字只用來事後核對、印警告，不參與計算）。

    回傳 CATEGORIES：[[label, start_no(1-based), end_no, total_price], ...]
    第一筆固定是「總計（全案）」涵蓋全部N_ITEMS個工項。
    """
    top_markers = []  # [start_item_no, numeral, label_text]
    mid_markers = []  # [start_item_no, numeral, label_text, top_index]
    item_no = 0
    cur_top_idx = None
    for rec in records:
        numeral = str(rec['項次']).strip() if not _is_blank(rec['項次']) else None
        if numeral in TOP_NUMERALS:
            top_markers.append([item_no + 1, numeral, rec['項目及說明']])
            cur_top_idx = len(top_markers) - 1
        elif numeral in MID_NUMERALS:
            mid_markers.append([item_no + 1, numeral, rec['項目及說明'], cur_top_idx])
        if rec['類型'] == '工項':
            item_no += 1

    for i in range(len(top_markers)):
        end = (top_markers[i + 1][0] - 1) if i + 1 < len(top_markers) else N_ITEMS
        top_markers[i].append(end)
    for i in range(len(mid_markers)):
        same_top_next = None
        for j in range(i + 1, len(mid_markers)):
            if mid_markers[j][3] == mid_markers[i][3]:
                same_top_next = mid_markers[j][0] - 1
                break
        if same_top_next is not None:
            end = same_top_next
        else:
            top_idx = mid_markers[i][3]
            end = top_markers[top_idx][3] if top_idx is not None else N_ITEMS
        mid_markers[i].append(end)

    CATEGORIES = [['總計（全案）', 1, N_ITEMS, grand_total]]
    for start, numeral, label, end in top_markers:
        text = str(label).split('\n')[0].split('\r')[0][:30] if label else ''
        total = round(sum(_item_amount(it) for it in items[start - 1:end]), 2)
        CATEGORIES.append([f'{numeral}、{text}' if text else numeral, start, end, total])
    for start, numeral, label, top_idx, end in mid_markers:
        text = str(label).split('\n')[0].split('\r')[0][:30] if label else ''
        total = round(sum(_item_amount(it) for it in items[start - 1:end]), 2)
        CATEGORIES.append([f'{numeral}、{text}' if text else numeral, start, end, total])

    check_total = sum(sum(_item_amount(it) for it in items[s - 1:e]) for _, s, e, _ in
                       [(m[1], m[0], m[3], None) for m in top_markers]) if top_markers else \
        sum(_item_amount(it) for it in items)
    if top_markers and abs(check_total - grand_total) > 1:
        print(f'⚠️ 警告：大項自算加總({check_total:,.0f}) 與 PCCES 總價({grand_total:,.0f}) 對不起來，'
              f'差額 {check_total - grand_total:,.0f}，請人工核對大項/中項編號是否有跳號或非標準寫法。')
    if not top_markers:
        print('⚠️ 提醒：偵測不到「壹貳參肆伍…」這種大項編號，此案將不會有大項/中項分類進度欄，'
              '但80個(或N個)工項本身仍會正常匯入、逐日填報功能不受影響。')
    return CATEGORIES


def parse_pcces_file(path, sheet_name=None, agency_override=None,
                      project_name_override=None, location_override=None,
                      proj_code_override=None):
    rows, used_sheet = read_pcces(path, sheet_name)
    header_row = find_header_row(rows)
    meta = extract_header_meta(rows, header_row)
    if project_name_override:
        meta['project_name'] = project_name_override
    if location_override:
        meta['location'] = location_override
    if proj_code_override:
        meta['proj_code'] = proj_code_override
    if agency_override:
        meta['agency'] = agency_override
    for k, v in meta.items():
        if v is None:
            print(f'⚠️ 提醒：偵測不到「{k}」，先留空，您可以在產生的活頁簿裡自行填寫，'
                  f'或下次執行時加 --{k.replace("_","-")} 參數手動指定。')
            meta[k] = ''

    records = parse_items(rows, header_row)
    items = [r for r in records if r['類型'] == '工項']
    N_ITEMS = len(items)
    if N_ITEMS == 0:
        raise ValueError('解析不到任何工項列，請確認分頁/檔案格式正確（用 --sheet 指定正確分頁）。')

    total_recs = [r for r in records if r['類型'] == '總計' and isinstance(r['複價'], (int, float))]
    self_sum = round(sum(_item_amount(it) for it in items), 2)
    plus_rec = next((r for r in total_recs if '+' in str(r['項目及說明'])), None)
    if plus_rec is not None:
        grand_total = plus_rec['複價']
        grand_total_source = f'PCCES檔案「{plus_rec["項目及說明"]}」列（含+號，判斷為最終總價）'
    elif total_recs:
        # 沒有明確列出「壹+貳+…」的加總列時，改信任文件裡最後一筆「總價」列（PCCES慣例上
        # 最後出現的總價列通常就是官方認定的契約金額；某些案子在這行之後才出現的項目，例如
        # 「剩餘價值折價費」這種扣款/調整性質的獨立項次，經查證屬於契約金額之外的另計項目，
        # 不應該被算進契約總價，所以優先信任官方標示的總價列，而不是把所有工項複價全部加總）。
        last_rec = total_recs[-1]
        grand_total = last_rec['複價']
        grand_total_source = f'PCCES檔案「{last_rec["項目及說明"]}」列（無+號，取文件中最後一筆總價列）'
    else:
        grand_total = self_sum
        grand_total_source = '偵測不到任何「總價」列，改用全部工項複價自行加總'
        print(f'⚠️ 提醒：偵測不到標準「總價」列，契約總價改用全部工項複價自行加總 = '
              f'{grand_total:,.0f}，請核對是否正確。')

    if abs(self_sum - grand_total) > 1:
        print(f'⚠️ 提醒：全部工項複價自行加總({self_sum:,.0f}) 與採用的契約總價'
              f'({grand_total:,.0f}，來源：{grand_total_source}) 不一致，差額 {self_sum - grand_total:,.0f}。'
              f'常見原因是文件裡「總價」列之後還有另計的扣款/調整項次（例如剩餘價值折價費），'
              f'不算入契約金額——請核對是否符合這種情況，或此案的「總價」列本身有跳號/非標準寫法。')

    CATEGORIES = build_categories(records, items, N_ITEMS, grand_total)
    print(f'工程名稱: {meta["project_name"]}　工項數量: {N_ITEMS}　契約總價: {grand_total:,.0f}'
          f'（來源：{grand_total_source}）　大項/中項數量: {len(CATEGORIES) - 1}　來源分頁: {used_sheet}')
    return meta, items, CATEGORIES, N_ITEMS, grand_total, records, grand_total_source, self_sum, used_sheet


# ============================================================
# 第一階段之附加功能：解析『變更設計比較表』格式
# ============================================================
# 使用者實測發現：實務上「套用契約變更」時拿到的檔案，往往不是全新的標準PCCES
# 「詳細價目表」匯出檔，而是機關/廠商慣用的「變更設計比較表」——同一列同時並列
# 「原契約」與「變更後」兩組數量/單價/複價，而且項次編號多了一層「壹/一/1/(1)」
# 四層結構(標準PCCES只有「壹/一」兩層)。這個附加功能讓 apply_change_order() 能夠
# 自動判斷收到的到底是哪一種格式，並且回傳跟 parse_pcces_file() 完全相同的
# tuple結構，讓下游(build_categories/apply_change_order本身)完全不用另外修改。
#
# 核心假設(已在下面各函式註解與呼叫端警告訊息裡列出，供使用者核對)：
# 1) 表頭「原契約」「變更後」兩組數量/單價欄位，一律以「左到右先出現的是原契約，
#    後出現的是變更後」判斷，這是這類比較表在實務上的通用排版慣例。
# 2) 大項(壹貳參...)/中項(一二三...)之間如果多了一層阿拉伯數字(1,2,3...)編號的
#    群組，一律併入所屬中項一起統計(不建立獨立的第三層進度)。
# 3) 「新增工項」(原契約數量為0/空、變更後數量>0)的單價/複價，如果檔案裡另外有
#    「議價明細表」(表頭含「需議價」字樣的分頁)，優先採用議價後的正式單價/複價；
#    找不到對應議價資料的新增工項，才退回使用變更比較表本身列出的初版單價。


def _norm_header_text(v):
    if not isinstance(v, str):
        return ''
    return (v.replace(' ', '').replace('　', '').replace('\n', '')
             .replace('(元)', '').replace('（元）', '').strip())


def _read_all_sheets_rows_xlsx(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    return [(ws.title, [[c.value for c in row] for row in ws.iter_rows()]) for ws in wb.worksheets]


def _read_all_sheets_rows_xls(path):
    import xlrd
    wb = xlrd.open_workbook(path)
    out = []
    for name in wb.sheet_names():
        sh = wb.sheet_by_name(name)
        out.append((name, [[sh.cell_value(r, c) for c in range(sh.ncols)] for r in range(sh.nrows)]))
    return out


def _read_all_sheets_rows(path):
    ext = os.path.splitext(path)[1].lower()
    if ext == '.xls':
        return _read_all_sheets_rows_xls(path)
    elif ext in ('.xlsx', '.xlsm'):
        return _read_all_sheets_rows_xlsx(path)
    else:
        raise ValueError(f'不支援的副檔名：{ext}（只支援 .xls / .xlsx / .xlsm）')


def _detect_change_comparison_header(rows, header_row):
    """判斷 header_row 是不是『變更設計比較表』格式的表頭：真正的判斷依據是
    header_row+1(子標題列)裡「數量」「單價」這兩種標籤各自出現兩次以上
    (分別對應「原契約」與「變更後」兩組欄位)，這跟標準PCCES「詳細價目表」
    (數量/單價各只出現一次)明顯不同，用這個當格式指紋，不依賴分頁名稱。"""
    if header_row + 1 >= len(rows):
        return False
    sub = rows[header_row + 1]
    qty_cols = [c for c, v in enumerate(sub) if _norm_header_text(v) == '數量']
    price_cols = [c for c, v in enumerate(sub) if _norm_header_text(v) == '單價']
    return len(qty_cols) >= 2 and len(price_cols) >= 2


def _find_change_comparison_columns(rows, header_row):
    """在header_row(項次/項目說明/單位等主標籤)與header_row+1(原契約/變更後的
    數量/單價/複價子標籤)兩列裡，用文字比對動態找出每個欄位的欄位編號——不假設
    固定欄位字母，因為不同案子/不同版本的變更比較表欄位位置不一定一樣。"""
    top = rows[header_row]
    sub = rows[header_row + 1] if header_row + 1 < len(rows) else []

    def find_in(row, texts):
        for c, v in enumerate(row):
            if _norm_header_text(v) in texts:
                return c
        return None

    item_no_col = find_in(top, {'項次'})
    desc_col = find_in(top, {'項目說明', '工程項目', '項目及說明', '項目', '說明', '項目名稱'})
    unit_col = find_in(top, {'單位'})
    if unit_col is None:
        unit_col = find_in(sub, {'單位'})
    note_col = find_in(top, {'備註'})
    if note_col is None:
        note_col = find_in(sub, {'備註'})

    qty_cols = [c for c, v in enumerate(sub) if _norm_header_text(v) == '數量']
    price_cols = [c for c, v in enumerate(sub) if _norm_header_text(v) == '單價']
    amount_cols = [c for c, v in enumerate(sub) if _norm_header_text(v) == '複價']

    if len(qty_cols) < 2 or len(price_cols) < 2:
        raise ValueError('偵測到疑似「變更設計比較表」格式，但找不到兩組完整的「數量/單價」欄位標籤，'
                          '無法自動判斷欄位位置。')
    missing = [name for name, col in [('項次', item_no_col), ('項目說明', desc_col), ('單位', unit_col)]
               if col is None]
    if missing:
        raise ValueError(f'「變更設計比較表」格式解析失敗：表頭裡找不到「{"、".join(missing)}」欄位標籤，'
                          f'請確認檔案格式或改用標準PCCES格式重新匯出。')

    orig_qty_col, new_qty_col = qty_cols[0], qty_cols[1]
    orig_price_col, new_price_col = price_cols[0], price_cols[1]
    orig_amount_col = amount_cols[0] if len(amount_cols) >= 1 else None
    new_amount_col = amount_cols[1] if len(amount_cols) >= 2 else None

    return dict(item_no_col=item_no_col, desc_col=desc_col, unit_col=unit_col, note_col=note_col,
                orig_qty_col=orig_qty_col, orig_price_col=orig_price_col, orig_amount_col=orig_amount_col,
                new_qty_col=new_qty_col, new_price_col=new_price_col, new_amount_col=new_amount_col)


def _parse_change_comparison_items(rows, header_row, col_map):
    """逐列讀取『變更設計比較表』，一律採用「變更後」欄位當作最新的數量/單價/
    複價(這是套用契約變更的核心目的)，並標記每一列是不是「新增工項」(原契約
    數量為0/空、變更後數量不為0)，供後續議價比對使用。"""
    records = []
    ncols = max(len(r) for r in rows) if rows else 1

    def col(row, key):
        idx = col_map[key]
        return row[idx] if idx is not None else None

    for r in range(header_row + 2, len(rows)):
        row = rows[r] + [None] * (ncols - len(rows[r]))
        item_no, desc, unit = col(row, 'item_no_col'), col(row, 'desc_col'), col(row, 'unit_col')
        new_qty, new_price = col(row, 'new_qty_col'), col(row, 'new_price_col')
        new_amount = col(row, 'new_amount_col')
        orig_qty = col(row, 'orig_qty_col')
        note = col(row, 'note_col')

        if all(_is_blank(x) for x in (item_no, desc, unit, new_qty, new_price, new_amount)):
            continue
        if _is_blank(desc):
            continue

        if _is_blank(new_amount) and not _is_blank(new_qty) and not _is_blank(new_price):
            try:
                new_amount = round(float(new_qty) * float(new_price), 2)
            except (TypeError, ValueError):
                pass

        orig_qty_num = orig_qty if isinstance(orig_qty, (int, float)) else None
        new_qty_num = new_qty if isinstance(new_qty, (int, float)) else None
        is_new_item = (orig_qty_num is None or orig_qty_num == 0) and (new_qty_num not in (None, 0))

        rec = {'項次': _fmt(item_no), '項目及說明': _fmt(desc), '單位': _fmt(unit),
               '數量': _fmt(new_qty), '單價': _fmt(new_price), '複價': _fmt(new_amount),
               '編碼備註': _fmt(note), '_is_new_item': is_new_item}
        records.append(rec)

    def classify(rec):
        text = str(rec['項目及說明'])
        has_qty = (not _is_blank(rec['單位'])) and (not _is_blank(rec['數量'])) and (not _is_blank(rec['單價']))
        if text.startswith('總價') or text.startswith('總計'):
            return '總計'
        if text.startswith('小計'):
            return '小計'
        if has_qty:
            return '工項'
        return '大項/中項'

    for rec in records:
        rec['類型'] = classify(rec)
    return records


def _parse_negotiated_prices(path):
    """掃描所有分頁，找出『新增工項議價明細表』格式的分頁(表頭含「需議價」字樣)，
    回傳 {項目說明文字: (議價後單價, 議價後複價)}。找不到就回傳空字典——代表這個
    案子沒有附議價明細表，呼叫端會維持用變更比較表自己列出的初版單價。"""
    try:
        raw_sheets = _read_all_sheets_rows(path)
    except Exception:
        return {}

    result = {}
    for name, rows in raw_sheets:
        header_row = None
        for r in range(min(15, len(rows))):
            row = rows[r]
            if not row:
                continue
            v0 = row[0]
            if isinstance(v0, str) and v0.strip() in ('項次', '項 次') and \
                    any('需議價' in _norm_header_text(v) for v in row):
                header_row = r
                break
        if header_row is None:
            continue

        top = rows[header_row]

        def find_in(row, texts):
            for c, v in enumerate(row):
                if _norm_header_text(v) in texts:
                    return c
            return None

        desc_col = find_in(top, {'項目及說明', '項目說明', '項目'})
        price_col = find_in(top, {'需議價單價'})
        amount_col = find_in(top, {'需議價複價'})
        if desc_col is None or price_col is None:
            continue

        ncols = max(len(r) for r in rows) if rows else 1
        for r in range(header_row + 1, len(rows)):
            row = rows[r] + [None] * (ncols - len(rows[r]))
            desc, price = row[desc_col], row[price_col]
            amount = row[amount_col] if amount_col is not None else None
            if _is_blank(desc) or _is_blank(price) or _is_blank(amount):
                continue
            result[str(desc).strip()] = (price, amount)
    return result


def looks_like_change_comparison_file(path, sheet_name=None):
    """快速判斷一份檔案裡是否存在『變更設計比較表』格式的分頁，給
    apply_change_order() 用來自動選擇該用哪一種解析器，不需要使用者手動指定。"""
    try:
        raw_sheets = _read_all_sheets_rows(path)
    except Exception:
        return False
    candidates = [(n, r) for n, r in raw_sheets if (sheet_name is None or n == sheet_name)]
    for _name, rows in candidates:
        try:
            hr = find_header_row(rows)
        except ValueError:
            continue
        if _detect_change_comparison_header(rows, hr):
            return True
    return False


def parse_change_comparison_file(path, sheet_name=None, agency_override=None,
                                  project_name_override=None, location_override=None,
                                  proj_code_override=None):
    """解析『變更設計比較表』格式(壹/一/阿拉伯數字群組/(細項)四層結構，同一列
    同時並列「原契約」與「變更後」兩組數量/單價/複價)，跟標準PCCES『詳細價目表』
    匯出格式完全不同。回傳格式跟 parse_pcces_file() 完全相同，方便
    apply_change_order() 等下游邏輯共用，不用另外改寫。

    如果檔案裡有多個分頁都符合這個格式的表頭指紋(常見情況：一份「N變更總表」的
    大項/中項彙總表 + 一份「N+1變更明細表」的完整逐項明細表都會符合)，一律採用
    「能解析出最多工項列」的那個分頁，因為明細表通常才是完整的工項清單。
    """
    raw_sheets = _read_all_sheets_rows(path)
    sheet_lookup = dict(raw_sheets)
    candidates = [sheet_name] if sheet_name else [name for name, _ in raw_sheets]

    best = None  # (n_items, sheet_name, rows, header_row, col_map, records)
    for name in candidates:
        rows = sheet_lookup.get(name)
        if rows is None:
            continue
        try:
            hr = find_header_row(rows)
        except ValueError:
            continue
        if not _detect_change_comparison_header(rows, hr):
            continue
        try:
            col_map = _find_change_comparison_columns(rows, hr)
            recs = _parse_change_comparison_items(rows, hr, col_map)
        except ValueError:
            continue
        n_items = sum(1 for r in recs if r['類型'] == '工項')
        if best is None or n_items > best[0]:
            best = (n_items, name, rows, hr, col_map, recs)

    if best is None:
        raise ValueError('偵測不到「變更設計比較表」格式的分頁(需要同一列同時出現兩組「數量/單價」欄位標籤)，'
                          '請確認檔案內容，或改用標準PCCES詳細價目表格式。')

    n_items, used_sheet, rows, header_row, col_map, records = best
    meta = extract_header_meta(rows, header_row)
    if project_name_override:
        meta['project_name'] = project_name_override
    if location_override:
        meta['location'] = location_override
    if proj_code_override:
        meta['proj_code'] = proj_code_override
    if agency_override:
        meta['agency'] = agency_override
    for k, v in meta.items():
        if v is None:
            print(f'⚠️ 提醒：偵測不到「{k}」，先留空，您可以在產生的活頁簿裡自行填寫。')
            meta[k] = ''

    items = [r for r in records if r['類型'] == '工項']
    N_ITEMS = len(items)
    if N_ITEMS == 0:
        raise ValueError('解析不到任何工項列，請確認分頁/檔案格式正確。')

    self_sum_before_negotiation = round(sum(_item_amount(it) for it in items), 2)

    negotiated = _parse_negotiated_prices(path)
    n_overridden = 0
    n_new_no_match = 0
    for it in items:
        if it.get('_is_new_item'):
            key = str(it['項目及說明']).strip()
            if key in negotiated:
                it['單價'], it['複價'] = negotiated[key]
                n_overridden += 1
            else:
                n_new_no_match += 1
    if n_overridden:
        print(f'ℹ️ {n_overridden} 個新增工項已採用「議價明細表」議價後的正式單價/複價'
              f'（優先於變更比較表本身列出的初版單價）。')
    if n_new_no_match:
        print(f'⚠️ 提醒：{n_new_no_match} 個新增工項找不到對應的議價明細表資料，仍採用變更比較表'
              f'本身列出的單價，建議人工核對是否已完成議價。')

    total_recs = [r for r in records if r['類型'] == '總計' and isinstance(r['複價'], (int, float))]
    self_sum = round(sum(_item_amount(it) for it in items), 2)
    plus_rec = next((r for r in total_recs if '+' in str(r['項目及說明'])), None)
    if plus_rec is not None:
        doc_total = plus_rec['複價']
        doc_total_source = f'變更比較表「{plus_rec["項目及說明"]}」列（含+號，判斷為最終總價）'
    elif total_recs:
        last_rec = total_recs[-1]
        doc_total = last_rec['複價']
        doc_total_source = f'變更比較表「{last_rec["項目及說明"]}」列（無+號，取文件中最後一筆總計列）'
    else:
        doc_total = None
        doc_total_source = None

    if doc_total is None:
        grand_total = self_sum
        grand_total_source = '偵測不到任何「總計/總價」列，改用全部工項複價自行加總'
        print(f'⚠️ 提醒：偵測不到標準「總計」列，契約總價改用全部工項複價自行加總 = '
              f'{grand_total:,.0f}，請核對是否正確。')
    else:
        # 沿用標準PCCES解析(parse_pcces_file)同一套既有慣例：優先信任文件本身標示的
        # 「總計」列，而不是把所有工項複價自行加總——因為「總計」列常常會刻意排除
        # 賸餘價值折價費之類的扣款/調整項次(這些不算入契約金額)，自行加總反而會算錯。
        grand_total = doc_total
        grand_total_source = doc_total_source

    negotiation_delta = round(self_sum - self_sum_before_negotiation, 2)
    if n_overridden and abs(negotiation_delta) > 1:
        print(f'ℹ️ 議價後正式單價讓新增工項的複價合計變動了 {negotiation_delta:,.0f} 元'
              f'（議價前自算加總{self_sum_before_negotiation:,.0f} → 議價後自算加總{self_sum:,.0f}）。'
              f'變更比較表本身的「總計」列({doc_total if doc_total is not None else 0:,.0f})多半是議價'
              f'「之前」就寫定的數字，可能沒有反映這筆變動，請人工核對「契約總價」欄位是否需要手動'
              f'改成議價後的正確金額。')

    if abs(self_sum - grand_total) > 1:
        print(f'⚠️ 提醒：全部工項複價自行加總({self_sum:,.0f}) 與採用的契約總價'
              f'({grand_total:,.0f}，來源：{grand_total_source}) 不一致，差額 {self_sum - grand_total:,.0f}，'
              f'常見原因是文件裡「總計」列之後還有另計的扣款/調整項次(例如賸餘價值折價費)，'
              f'或是議價後正式單價還沒反映進文件的「總計」列(見上方訊息)，請核對是否符合其中一種情況。')

    CATEGORIES = build_categories(records, items, N_ITEMS, grand_total)
    print(f'【變更設計比較表格式】工程名稱: {meta["project_name"]}　工項數量: {N_ITEMS}　'
          f'變更後契約總價: {grand_total:,.0f}（來源：{grand_total_source}）　'
          f'大項/中項數量: {len(CATEGORIES) - 1}　來源分頁: {used_sheet}')
    return meta, items, CATEGORIES, N_ITEMS, grand_total, records, grand_total_source, self_sum, used_sheet


# ============================================================
# 第二階段：日資料庫 / 預定進度 / 甘特圖
# ============================================================

def build_day_database(wb, meta, items, CATEGORIES, N_ITEMS, grand_total,
                        start_date, days, n_mat, n_labor, n_mach):
    ws = wb.active
    ws.title = '日資料庫'

    thin = Side(style='thin', color='BFBFBF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill('solid', fgColor='D9E1F2')
    cat_fill = PatternFill('solid', fgColor='FCE4D6')
    info_fill = PatternFill('solid', fgColor='FFF2CC')
    mat_fill = PatternFill('solid', fgColor='E2EFDA')
    txt_fill = PatternFill('solid', fgColor='EDEDED')
    bold = Font(bold=True)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws['A1'] = '工程名稱'; ws['A1'].font = bold
    ws['B1'] = meta['project_name']; ws.merge_cells('B1:E1')
    ws['F1'] = '工程編號'; ws['F1'].font = bold
    ws['G1'] = meta['proj_code']
    ws['H1'] = '契約總價'; ws['H1'].font = bold
    ws['I1'] = grand_total; ws['I1'].number_format = '#,##0'
    ws['J1'] = '主辦機關'; ws['J1'].font = bold
    ws['K1'] = meta['agency']
    ws['L1'] = '工期展延天數'; ws['L1'].font = bold
    ws['M1'] = 0; ws['M1'].fill = info_fill

    ws['A2'] = '開工日期'; ws['A2'].font = bold
    ws['B2'] = start_date; ws['B2'].number_format = 'yyyy/mm/dd'; ws['B2'].fill = info_fill
    ws['C2'] = '工期(日曆天)'; ws['C2'].font = bold
    ws['D2'] = days; ws['D2'].fill = info_fill
    ws['E2'] = '預定完工日'; ws['E2'].font = bold
    ws['F2'] = '=B2+D2-1'; ws['F2'].number_format = 'yyyy/mm/dd'
    ws['G2'] = '施工地點'; ws['G2'].font = bold
    ws['H2'] = meta['location']
    note = ws.cell(row=2, column=9, value='← 請將 B2 改為實際開工日期，全部日期會自動跟著更新')
    note.font = Font(italic=True, color='C00000')

    ROW_SEQ, ROW_NAME, ROW_UNIT, ROW_QTY, ROW_PRICE, ROW_LABEL = 4, 5, 6, 7, 8, 9
    ROW_REMARK = 3
    FIRST_DAY_ROW = 10
    LAST_DAY_ROW = FIRST_DAY_ROW + days - 1

    labels_ABCDE = ['第幾天', '日期', '星期', '天氣(上午)', '天氣(下午)']
    for j, v in enumerate(labels_ABCDE, start=1):
        c = ws.cell(row=ROW_NAME, column=j, value=v)
        c.font = bold; c.fill = head_fill; c.alignment = center; c.border = border
    for row_idx in (ROW_SEQ, ROW_UNIT, ROW_QTY, ROW_PRICE, ROW_LABEL):
        for j in range(1, 6):
            ws.cell(row=row_idx, column=j).border = border
    ws.cell(row=ROW_LABEL, column=1, value='(每日填寫)').font = Font(italic=True)

    FIRST_CAT_COL = 6
    LAST_CAT_COL = FIRST_CAT_COL + len(CATEGORIES) - 1
    FIRST_ITEM_COL = LAST_CAT_COL + 1
    LAST_ITEM_COL = FIRST_ITEM_COL + N_ITEMS - 1
    FIRST_MAT_COL = LAST_ITEM_COL + 1
    LAST_MAT_COL = FIRST_MAT_COL + n_mat - 1
    FIRST_LAB_COL = LAST_MAT_COL + 1
    LAST_LAB_COL = FIRST_LAB_COL + n_labor - 1
    FIRST_MCH_COL = LAST_LAB_COL + 1
    LAST_MCH_COL = FIRST_MCH_COL + n_mach - 1

    # TXT_FIELDS 現在是模組層級常數（見檔案上方），這裡不再重複定義。
    FIRST_TXT_COL = LAST_MCH_COL + 1
    LAST_TXT_COL = FIRST_TXT_COL + len(TXT_FIELDS) - 1

    N_SCHED = 30
    SCHED_SHEET_NAME = '預定進度'
    SCHED_TASK_FIRST_COL = 3
    SCHED_TASK_LAST_COL = SCHED_TASK_FIRST_COL + N_SCHED - 1
    SCHED_TOTAL_COL = SCHED_TASK_LAST_COL + 1
    SCHED_TOTAL_COL_LETTER = get_column_letter(SCHED_TOTAL_COL)
    SCHED_ROW_NAME, SCHED_ROW_AMOUNT, SCHED_ROW_WEIGHT, SCHED_ROW_START, SCHED_ROW_END, SCHED_ROW_LABEL = 4, 5, 6, 7, 8, 9

    item_col_letters = [get_column_letter(FIRST_ITEM_COL + i) for i in range(N_ITEMS)]
    new_categories = []
    for ci, (label, start_no, end_no, total_price) in enumerate(CATEGORIES):
        col = FIRST_CAT_COL + ci
        col_letter = get_column_letter(col)
        c1 = item_col_letters[start_no - 1]
        c2 = item_col_letters[end_no - 1]
        cells = {ROW_SEQ: '', ROW_NAME: label, ROW_UNIT: '進度%', ROW_QTY: '',
                 ROW_PRICE: total_price, ROW_LABEL: '累計完成率(自動)'}
        for row_idx, val in cells.items():
            c = ws.cell(row=row_idx, column=col, value=val)
            c.border = border; c.fill = cat_fill
            c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True) if row_idx == ROW_NAME else center
            if row_idx == ROW_PRICE and isinstance(val, (int, float)):
                c.number_format = '#,##0'
        new_categories.append((label, start_no, end_no, total_price, col_letter, c1, c2))
    CATEGORIES_full = new_categories

    # 中項欄位預設摺疊（用Excel大綱分組功能，資料完全還在、只是先收起來省畫面空間）：
    # CATEGORIES 的組成順序固定是「總計＋全部大項＋全部中項」（見 build_categories()），
    # 所以中項一定是最後連續一段，可以直接整段分組，不需要逐欄判斷是否相鄰。
    mid_level_cols = [col_letter for (label, sn, en, tp, col_letter, c1, c2) in CATEGORIES_full
                       if label and label[0] in MID_NUMERALS]
    if mid_level_cols:
        ws.sheet_properties.outlinePr = Outline(summaryRight=False)
        for col_letter in mid_level_cols:
            cdim = ws.column_dimensions[col_letter]
            cdim.outlineLevel = 1
            cdim.hidden = True
        n_visible_cat = len(CATEGORIES) - len(mid_level_cols)
        note_col = FIRST_CAT_COL + n_visible_cat - 1
        note_cat = ws.cell(row=ROW_SEQ, column=note_col,
                            value=f'※中項欄位（{len(mid_level_cols)}欄，緊接在右側）預設已摺疊以節省畫面空間；'
                                  f'點選欄號上方的「+」即可展開查看，資料都還在，摺疊只是視覺收合。')
        note_cat.font = Font(italic=True, color='C00000', size=8)

    for idx, item in enumerate(items):
        col = FIRST_ITEM_COL + idx
        cells = {ROW_SEQ: idx + 1, ROW_NAME: item['項目及說明'], ROW_UNIT: item['單位'],
                 ROW_QTY: item['數量'], ROW_PRICE: item['單價'], ROW_LABEL: '本日完成數量',
                 ROW_REMARK: ''}
        for row_idx, val in cells.items():
            c = ws.cell(row=row_idx, column=col, value=val)
            c.border = border; c.fill = head_fill
            c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True) if row_idx == ROW_NAME else center
            if row_idx in (ROW_QTY, ROW_PRICE) and isinstance(val, (int, float)):
                c.number_format = '#,##0.##'

    for i in range(n_mat):
        col = FIRST_MAT_COL + i
        cells = {ROW_SEQ: '', ROW_NAME: '', ROW_UNIT: '', ROW_QTY: '', ROW_PRICE: '', ROW_LABEL: '本日使用數量', ROW_REMARK: ''}
        for row_idx, val in cells.items():
            c = ws.cell(row=row_idx, column=col, value=val)
            c.border = border; c.fill = mat_fill
            c.alignment = center
    note_mat = ws.cell(row=ROW_SEQ, column=FIRST_MAT_COL,
                        value=f'※材料{n_mat}筆：名稱/單位/契約數量/備註請填在下方第{ROW_NAME}~{ROW_QTY}及第{ROW_REMARK}列，不夠可自行往右插入欄位')
    note_mat.font = Font(italic=True, color='C00000', size=8)
    ws.merge_cells(start_row=ROW_SEQ, start_column=FIRST_MAT_COL, end_row=ROW_SEQ, end_column=min(LAST_MAT_COL, FIRST_MAT_COL + 3))

    for i in range(n_labor):
        col = FIRST_LAB_COL + i
        name = DEFAULT_LABOR_NAMES[i] if i < len(DEFAULT_LABOR_NAMES) else ''
        unit = DEFAULT_LABOR_UNIT if name else ''
        cells = {ROW_SEQ: '', ROW_NAME: name, ROW_UNIT: unit, ROW_QTY: '', ROW_PRICE: '', ROW_LABEL: '本日人數'}
        for row_idx, val in cells.items():
            c = ws.cell(row=row_idx, column=col, value=val)
            c.border = border; c.fill = mat_fill
            c.alignment = center
    note_lab = ws.cell(row=ROW_SEQ, column=FIRST_LAB_COL,
                        value=f'※工別{n_labor}筆：已預填{len(DEFAULT_LABOR_NAMES)}個常用名稱(單位預設「工」)，'
                              f'可直接覆寫成貴案實際工別，超過的空格請自行增填')
    note_lab.font = Font(italic=True, color='C00000', size=8)
    ws.merge_cells(start_row=ROW_SEQ, start_column=FIRST_LAB_COL, end_row=ROW_SEQ, end_column=min(LAST_LAB_COL, FIRST_LAB_COL + 3))

    for i in range(n_mach):
        col = FIRST_MCH_COL + i
        name = DEFAULT_MACH_NAMES[i] if i < len(DEFAULT_MACH_NAMES) else ''
        unit = DEFAULT_MACH_UNIT if name else ''
        cells = {ROW_SEQ: '', ROW_NAME: name, ROW_UNIT: unit, ROW_QTY: '', ROW_PRICE: '', ROW_LABEL: '本日使用數量'}
        for row_idx, val in cells.items():
            c = ws.cell(row=row_idx, column=col, value=val)
            c.border = border; c.fill = mat_fill
            c.alignment = center
    note_mch = ws.cell(row=ROW_SEQ, column=FIRST_MCH_COL,
                        value=f'※機具{n_mach}筆：已預填{len(DEFAULT_MACH_NAMES)}個常用名稱(單位預設「台班」)，'
                              f'可直接覆寫成貴案實際機具，超過的空格請自行增填')
    note_mch.font = Font(italic=True, color='C00000', size=8)
    ws.merge_cells(start_row=ROW_SEQ, start_column=FIRST_MCH_COL, end_row=ROW_SEQ, end_column=min(LAST_MCH_COL, FIRST_MCH_COL + 3))

    dv_yn = DataValidation(type='list', formula1='"有,無"', allow_blank=True)
    dv_yn2 = DataValidation(type='list', formula1='"有,無,無新進勞工"', allow_blank=True)
    ws.add_data_validation(dv_yn)
    ws.add_data_validation(dv_yn2)

    for idx, (key, full_text, kind) in enumerate(TXT_FIELDS):
        col = FIRST_TXT_COL + idx
        for row_idx in (ROW_SEQ, ROW_NAME, ROW_UNIT, ROW_QTY, ROW_PRICE, ROW_LABEL):
            c = ws.cell(row=row_idx, column=col); c.fill = txt_fill; c.border = border
        ws.cell(row=ROW_NAME, column=col, value=full_text).font = bold
        ws.cell(row=ROW_NAME, column=col).alignment = center

    for d in range(days):
        r = FIRST_DAY_ROW + d
        ws.cell(row=r, column=1, value=d + 1)
        ws.cell(row=r, column=2, value='=$B$2' if d == 0 else f'=B{r-1}+1')
        ws.cell(row=r, column=2).number_format = 'yyyy/mm/dd'
        ws.cell(row=r, column=3, value=f'=CHOOSE(WEEKDAY(B{r},2),"星期一","星期二","星期三","星期四","星期五","星期六","星期日")')
        for _, _, _, _, col_letter, c1, c2 in CATEGORIES_full:
            today = f"SUMPRODUCT(${c1}${ROW_PRICE}:${c2}${ROW_PRICE},{c1}{r}:{c2}{r})/${col_letter}${ROW_PRICE}"
            f = f"=IFERROR({today},0)" if d == 0 else f"=IFERROR({today},0)+{col_letter}{r-1}"
            cell = ws.cell(row=r, column=openpyxl.utils.column_index_from_string(col_letter), value=f)
            cell.number_format = '0.00%'
        for col in range(1, LAST_TXT_COL + 1):
            c = ws.cell(row=r, column=col)
            c.border = border
            c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True) if col >= FIRST_TXT_COL else Alignment(horizontal='center', vertical='center')
        sched_col = FIRST_TXT_COL + [i for i, (k, _, _) in enumerate(TXT_FIELDS) if k == 'SCHED'][0]
        sc = ws.cell(row=r, column=sched_col, value=f"='{SCHED_SHEET_NAME}'!{SCHED_TOTAL_COL_LETTER}{r}")
        sc.number_format = '0.00%'
        sc.alignment = Alignment(horizontal='center', vertical='center')

    for idx, (key, full_text, kind) in enumerate(TXT_FIELDS):
        col = FIRST_TXT_COL + idx
        if kind == 'yn':
            for d in range(days):
                dv_yn.add(ws.cell(row=FIRST_DAY_ROW + d, column=col))
        elif kind == 'yn2':
            for d in range(days):
                dv_yn2.add(ws.cell(row=FIRST_DAY_ROW + d, column=col))

    dv_weather = DataValidation(type='list', formula1='"晴,陰,雨,颱風"', allow_blank=True)
    ws.add_data_validation(dv_weather)
    for d in range(days):
        r = FIRST_DAY_ROW + d
        dv_weather.add(ws.cell(row=r, column=4))
        dv_weather.add(ws.cell(row=r, column=5))

    ws.freeze_panes = ws.cell(row=FIRST_DAY_ROW, column=FIRST_ITEM_COL)
    ws.column_dimensions['A'].width = 7
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    for col in range(FIRST_CAT_COL, LAST_CAT_COL + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16
    for col in range(FIRST_ITEM_COL, LAST_ITEM_COL + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14
    for col in range(FIRST_MAT_COL, LAST_MCH_COL + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14
    for col in range(FIRST_TXT_COL, LAST_TXT_COL + 1):
        ws.column_dimensions[get_column_letter(col)].width = 30
    ws.row_dimensions[ROW_NAME].height = 60
    ws.sheet_view.showGridLines = False
    note2 = ws.cell(row=3, column=1, value='※本工作表為每日輸入資料庫，請至「施工日報」工作表查詢/列印當日報表')
    note2.font = Font(italic=True, color='C00000')
    ws.merge_cells('A3:I3')
    ws.print_area = 'A1:I9'
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1

    ctx = dict(
        FIRST_CAT_COL=FIRST_CAT_COL, LAST_CAT_COL=LAST_CAT_COL,
        FIRST_ITEM_COL=FIRST_ITEM_COL, LAST_ITEM_COL=LAST_ITEM_COL, N_ITEMS=N_ITEMS,
        FIRST_MAT_COL=FIRST_MAT_COL, LAST_MAT_COL=LAST_MAT_COL, N_MAT=n_mat,
        FIRST_LAB_COL=FIRST_LAB_COL, LAST_LAB_COL=LAST_LAB_COL, N_LABOR=n_labor,
        FIRST_MCH_COL=FIRST_MCH_COL, LAST_MCH_COL=LAST_MCH_COL, N_MACH=n_mach,
        FIRST_TXT_COL=FIRST_TXT_COL, LAST_TXT_COL=LAST_TXT_COL, TXT_FIELDS=TXT_FIELDS,
        ROW_SEQ=ROW_SEQ, ROW_NAME=ROW_NAME, ROW_UNIT=ROW_UNIT, ROW_QTY=ROW_QTY,
        ROW_PRICE=ROW_PRICE, ROW_LABEL=ROW_LABEL, ROW_REMARK=ROW_REMARK,
        FIRST_DAY_ROW=FIRST_DAY_ROW, LAST_DAY_ROW=LAST_DAY_ROW, DAYS=days,
        CATEGORIES_full=CATEGORIES_full,
        SCHED_SHEET_NAME=SCHED_SHEET_NAME, N_SCHED=N_SCHED,
        SCHED_TASK_FIRST_COL=SCHED_TASK_FIRST_COL, SCHED_TASK_LAST_COL=SCHED_TASK_LAST_COL,
        SCHED_TOTAL_COL=SCHED_TOTAL_COL, SCHED_TOTAL_COL_LETTER=SCHED_TOTAL_COL_LETTER,
        SCHED_ROW_NAME=SCHED_ROW_NAME, SCHED_ROW_AMOUNT=SCHED_ROW_AMOUNT, SCHED_ROW_WEIGHT=SCHED_ROW_WEIGHT,
        SCHED_ROW_START=SCHED_ROW_START, SCHED_ROW_END=SCHED_ROW_END, SCHED_ROW_LABEL=SCHED_ROW_LABEL,
        start_date=start_date,
    )
    return ctx


def build_sched_and_gantt(wb, meta, ctx):
    bold = Font(bold=True)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='BFBFBF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill('solid', fgColor='D9E1F2')
    info_fill = PatternFill('solid', fgColor='FFF2CC')

    SCHED_SHEET_NAME = ctx['SCHED_SHEET_NAME']
    N_SCHED = ctx['N_SCHED']
    SCHED_TASK_FIRST_COL = ctx['SCHED_TASK_FIRST_COL']
    SCHED_TOTAL_COL = ctx['SCHED_TOTAL_COL']
    SCHED_TOTAL_COL_LETTER = ctx['SCHED_TOTAL_COL_LETTER']
    SCHED_ROW_NAME, SCHED_ROW_AMOUNT, SCHED_ROW_WEIGHT, SCHED_ROW_START, SCHED_ROW_END, SCHED_ROW_LABEL = (
        ctx['SCHED_ROW_NAME'], ctx['SCHED_ROW_AMOUNT'], ctx['SCHED_ROW_WEIGHT'],
        ctx['SCHED_ROW_START'], ctx['SCHED_ROW_END'], ctx['SCHED_ROW_LABEL'])
    FIRST_DAY_ROW, days = ctx['FIRST_DAY_ROW'], ctx['DAYS']
    start_date = ctx['start_date']

    ws2 = wb.create_sheet(SCHED_SHEET_NAME)
    ws2.sheet_view.showGridLines = False
    ws2['A1'] = '工程名稱'; ws2['A1'].font = bold
    ws2['B1'] = meta['project_name']; ws2.merge_cells('B1:E1')

    note_s = ws2.cell(row=2, column=1,
        value='※請在下方任務清單填入「任務名稱/契約金額/開始日期/結束日期」（可自行增列，本頁預留'
              f'{N_SCHED}個任務欄位）。本頁會假設每個任務在起訖日期間「逐日直線內插」完成度，用契約'
              '金額佔比當權重加權加總，自動算出全案逐日「預定累計進度%」，並自動回填至「日資料庫」'
              '的預定進度(%)欄，不需要在日資料庫手動輸入。如任務起訖日與實際排程有落差(例如你們有'
              '更精確的S曲線/核定進度表)，請改用那份資料的數字，不要用本頁反推。')
    note_s.font = Font(italic=True, color='C00000')
    ws2.merge_cells(start_row=2, start_column=1, end_row=2, end_column=SCHED_TOTAL_COL)
    ws2.row_dimensions[2].height = 30

    labels_sched = {SCHED_ROW_NAME: '任務名稱', SCHED_ROW_AMOUNT: '契約金額(權重來源)',
                    SCHED_ROW_WEIGHT: '權重%(自動)', SCHED_ROW_START: '開始日期',
                    SCHED_ROW_END: '結束日期'}
    for row_idx, text in labels_sched.items():
        c = ws2.cell(row=row_idx, column=2, value=text)
        c.font = bold; c.fill = head_fill; c.alignment = Alignment(horizontal='left', vertical='center')
        c.border = border
        ws2.cell(row=row_idx, column=1).border = border

    task_col_letters = [get_column_letter(SCHED_TASK_FIRST_COL + i) for i in range(N_SCHED)]
    weight_range = f"${task_col_letters[0]}${SCHED_ROW_AMOUNT}:${task_col_letters[-1]}${SCHED_ROW_AMOUNT}"

    for i, cl in enumerate(task_col_letters):
        col = SCHED_TASK_FIRST_COL + i
        ws2.cell(row=SCHED_ROW_NAME, column=col, value='').fill = info_fill
        amt = ws2.cell(row=SCHED_ROW_AMOUNT, column=col, value='')
        amt.fill = info_fill; amt.number_format = '#,##0'
        wgt = ws2.cell(row=SCHED_ROW_WEIGHT, column=col,
                        value=f"=IFERROR({cl}{SCHED_ROW_AMOUNT}/SUM({weight_range}),0)")
        wgt.number_format = '0.00%'
        st = ws2.cell(row=SCHED_ROW_START, column=col, value='')
        st.fill = info_fill; st.number_format = 'yyyy/mm/dd'
        en = ws2.cell(row=SCHED_ROW_END, column=col, value='')
        en.fill = info_fill; en.number_format = 'yyyy/mm/dd'
        for row_idx in (SCHED_ROW_NAME, SCHED_ROW_AMOUNT, SCHED_ROW_WEIGHT, SCHED_ROW_START, SCHED_ROW_END):
            ws2.cell(row=row_idx, column=col).border = border
        ws2.cell(row=SCHED_ROW_LABEL, column=col, value=f'任務{i+1}').font = Font(italic=True, size=9)
        ws2.cell(row=SCHED_ROW_LABEL, column=col).alignment = center
        ws2.cell(row=SCHED_ROW_LABEL, column=col).border = border

    demo_col = SCHED_TASK_FIRST_COL
    ws2.cell(row=SCHED_ROW_NAME, column=demo_col, value='(範例，請覆寫成實際任務)整地工程')
    ws2.cell(row=SCHED_ROW_AMOUNT, column=demo_col, value=1000000)
    ws2.cell(row=SCHED_ROW_START, column=demo_col, value=start_date)
    ws2.cell(row=SCHED_ROW_END, column=demo_col, value=start_date + datetime.timedelta(days=9))

    chk = ws2.cell(row=SCHED_ROW_WEIGHT, column=SCHED_TOTAL_COL,
                    value=f"=SUM({task_col_letters[0]}{SCHED_ROW_WEIGHT}:{task_col_letters[-1]}{SCHED_ROW_WEIGHT})")
    chk.number_format = '0.00%'; chk.font = bold
    ws2.cell(row=SCHED_ROW_NAME, column=SCHED_TOTAL_COL, value='權重合計(應=100%)').font = Font(italic=True, size=9)

    ws2.cell(row=SCHED_ROW_LABEL, column=1, value='第幾天').font = bold
    ws2.cell(row=SCHED_ROW_LABEL, column=1).fill = head_fill
    ws2.cell(row=SCHED_ROW_LABEL, column=1).border = border
    ws2.cell(row=SCHED_ROW_LABEL, column=2, value='日期').font = bold
    ws2.cell(row=SCHED_ROW_LABEL, column=2).fill = head_fill
    ws2.cell(row=SCHED_ROW_LABEL, column=2).border = border
    ws2.cell(row=SCHED_ROW_LABEL, column=SCHED_TOTAL_COL, value='預定累計進度%(自動)').font = bold
    ws2.cell(row=SCHED_ROW_LABEL, column=SCHED_TOTAL_COL).fill = head_fill
    ws2.cell(row=SCHED_ROW_LABEL, column=SCHED_TOTAL_COL).alignment = center
    ws2.cell(row=SCHED_ROW_LABEL, column=SCHED_TOTAL_COL).border = border

    for d in range(days):
        r = FIRST_DAY_ROW + d
        ws2.cell(row=r, column=1, value=d + 1)
        dcell = ws2.cell(row=r, column=2, value=f"='日資料庫'!B{r}")
        dcell.number_format = 'yyyy/mm/dd'
        for i, cl in enumerate(task_col_letters):
            col = SCHED_TASK_FIRST_COL + i
            frac = (f"=IF(OR({cl}${SCHED_ROW_START}=\"\",{cl}${SCHED_ROW_END}=\"\"),0,"
                    f"IF($B{r}<{cl}${SCHED_ROW_START},0,IF($B{r}>={cl}${SCHED_ROW_END},1,"
                    f"($B{r}-{cl}${SCHED_ROW_START}+1)/({cl}${SCHED_ROW_END}-{cl}${SCHED_ROW_START}+1))))")
            fc = ws2.cell(row=r, column=col, value=frac)
            fc.number_format = '0.00%'
        total_f = (f"=SUMPRODUCT({task_col_letters[0]}${SCHED_ROW_WEIGHT}:{task_col_letters[-1]}${SCHED_ROW_WEIGHT},"
                   f"{task_col_letters[0]}{r}:{task_col_letters[-1]}{r})")
        tc = ws2.cell(row=r, column=SCHED_TOTAL_COL, value=total_f)
        tc.number_format = '0.00%'; tc.font = bold
        for col in range(1, SCHED_TOTAL_COL + 1):
            ws2.cell(row=r, column=col).border = border

    ws2.freeze_panes = ws2.cell(row=FIRST_DAY_ROW, column=SCHED_TASK_FIRST_COL)
    ws2.column_dimensions['A'].width = 7
    ws2.column_dimensions['B'].width = 12
    for cl in task_col_letters:
        ws2.column_dimensions[cl].width = 12
    ws2.column_dimensions[SCHED_TOTAL_COL_LETTER].width = 16
    ws2.print_area = f'A1:{SCHED_TOTAL_COL_LETTER}9'
    ws2.page_setup.fitToPage = True
    ws2.sheet_properties.pageSetUpPr.fitToPage = True
    ws2.page_setup.fitToWidth = 1
    ws2.page_setup.fitToHeight = 1

    # ---- 甘特圖 ----
    GANTT_SHEET_NAME = '甘特圖'
    wsg = wb.create_sheet(GANTT_SHEET_NAME)
    wsg.sheet_view.showGridLines = False
    head_fill2 = PatternFill('solid', fgColor='D9E1F2')

    G_LABEL_COLS = 4
    G_DAY_FIRST_COL = G_LABEL_COLS + 1
    G_DAY_LAST_COL = G_DAY_FIRST_COL + days - 1
    G_ROW_HEADER = 2
    G_ROW_TASK_FIRST = 3
    G_ROW_TASK_LAST = G_ROW_TASK_FIRST + N_SCHED - 1
    G_ROW_CUM_SCHED = G_ROW_TASK_LAST + 2
    G_ROW_CUM_ACTUAL = G_ROW_CUM_SCHED + 1

    wsg['A1'] = f'{meta["project_name"]}　甘特圖(預定進度) — 資料來源：預定進度／日資料庫分頁，此頁不需手動輸入'
    wsg['A1'].font = Font(bold=True, size=12)

    label_headers = ['任務名稱', '開始日期', '結束日期', '權重%']
    for j, t in enumerate(label_headers, start=1):
        c = wsg.cell(row=G_ROW_HEADER, column=j, value=t)
        c.font = bold; c.fill = head_fill2; c.alignment = center; c.border = border

    day_font = Font(size=7)
    for d in range(days):
        col = G_DAY_FIRST_COL + d
        c = wsg.cell(row=G_ROW_HEADER, column=col, value=f"='日資料庫'!B{FIRST_DAY_ROW + d}")
        c.number_format = 'm/d'
        c.font = day_font
        c.alignment = Alignment(horizontal='center', vertical='center', text_rotation=90)
        c.border = border

    bar_fill = PatternFill('solid', fgColor='4472C4')
    last_day_letter = get_column_letter(G_DAY_LAST_COL)
    first_day_letter = get_column_letter(G_DAY_FIRST_COL)

    for i, cl in enumerate(task_col_letters):
        r = G_ROW_TASK_FIRST + i
        wsg.cell(row=r, column=1,
                 value=f"=IF('{SCHED_SHEET_NAME}'!{cl}{SCHED_ROW_NAME}=\"\",\"\",'{SCHED_SHEET_NAME}'!{cl}{SCHED_ROW_NAME})")
        st = wsg.cell(row=r, column=2,
                      value=f"=IF('{SCHED_SHEET_NAME}'!{cl}{SCHED_ROW_START}=\"\",\"\",'{SCHED_SHEET_NAME}'!{cl}{SCHED_ROW_START})")
        st.number_format = 'yyyy/mm/dd'
        en = wsg.cell(row=r, column=3,
                      value=f"=IF('{SCHED_SHEET_NAME}'!{cl}{SCHED_ROW_END}=\"\",\"\",'{SCHED_SHEET_NAME}'!{cl}{SCHED_ROW_END})")
        en.number_format = 'yyyy/mm/dd'
        wt = wsg.cell(row=r, column=4, value=f"='{SCHED_SHEET_NAME}'!{cl}{SCHED_ROW_WEIGHT}")
        wt.number_format = '0.00%'
        for col in range(1, 5):
            wsg.cell(row=r, column=col).border = border
        rng = f"{first_day_letter}{r}:{last_day_letter}{r}"
        formula = (f'AND({first_day_letter}${G_ROW_HEADER}<>"",$B{r}<>"",$C{r}<>"",'
                   f'{first_day_letter}${G_ROW_HEADER}>=$B{r},{first_day_letter}${G_ROW_HEADER}<=$C{r})')
        wsg.conditional_formatting.add(rng, FormulaRule(formula=[formula], fill=bar_fill))

    wsg.cell(row=G_ROW_CUM_SCHED, column=1, value='預定累計進度%(甘特圖加總)').font = bold
    wsg.cell(row=G_ROW_CUM_SCHED, column=4, value='預定累計進度%')
    wsg.cell(row=G_ROW_CUM_ACTUAL, column=1, value='實際累計進度%(對照用)').font = bold
    wsg.cell(row=G_ROW_CUM_ACTUAL, column=4, value='實際累計進度%')
    actual_col_letter = ctx['CATEGORIES_full'][0][4]
    for d in range(days):
        col = G_DAY_FIRST_COL + d
        r_day = FIRST_DAY_ROW + d
        sc = wsg.cell(row=G_ROW_CUM_SCHED, column=col, value=f"='{SCHED_SHEET_NAME}'!{SCHED_TOTAL_COL_LETTER}{r_day}")
        sc.number_format = '0%'; sc.font = Font(size=7)
        ac = wsg.cell(row=G_ROW_CUM_ACTUAL, column=col, value=f"='日資料庫'!{actual_col_letter}{r_day}")
        ac.number_format = '0%'; ac.font = Font(size=7)

    wsg.freeze_panes = wsg.cell(row=G_ROW_TASK_FIRST, column=G_DAY_FIRST_COL)
    wsg.column_dimensions['A'].width = 20
    for c in ('B', 'C', 'D'):
        wsg.column_dimensions[c].width = 11
    for d in range(days):
        wsg.column_dimensions[get_column_letter(G_DAY_FIRST_COL + d)].width = 2.6
    wsg.row_dimensions[G_ROW_HEADER].height = 40

    chart = LineChart()
    chart.title = '預定 vs 實際 累計進度 S曲線'
    chart.y_axis.title = '累計進度%'
    chart.x_axis.title = '日期'
    chart.height = 8
    chart.width = 30
    data = Reference(wsg, min_col=4, max_col=G_DAY_LAST_COL, min_row=G_ROW_CUM_SCHED, max_row=G_ROW_CUM_ACTUAL)
    chart.add_data(data, titles_from_data=True, from_rows=True)
    cats = Reference(wsg, min_col=G_DAY_FIRST_COL, max_col=G_DAY_LAST_COL, min_row=G_ROW_HEADER, max_row=G_ROW_HEADER)
    chart.set_categories(cats)
    chart_anchor_row = G_ROW_CUM_ACTUAL + 3
    wsg.add_chart(chart, f"A{chart_anchor_row}")

    chart_col_span = 46
    chart_row_span = 18
    print_last_col_letter = get_column_letter(chart_col_span)
    wsg.print_area = f"A{chart_anchor_row}:{print_last_col_letter}{chart_anchor_row + chart_row_span}"
    wsg.page_setup.orientation = 'landscape'
    wsg.page_setup.fitToPage = True
    wsg.sheet_properties.pageSetUpPr.fitToPage = True
    wsg.page_setup.fitToWidth = 1
    wsg.page_setup.fitToHeight = 1

    ctx.update(dict(
        GANTT_SHEET_NAME=GANTT_SHEET_NAME, G_DAY_FIRST_COL=G_DAY_FIRST_COL, G_DAY_LAST_COL=G_DAY_LAST_COL,
        G_ROW_TASK_FIRST=G_ROW_TASK_FIRST, G_ROW_TASK_LAST=G_ROW_TASK_LAST,
        G_ROW_CUM_SCHED=G_ROW_CUM_SCHED, G_ROW_CUM_ACTUAL=G_ROW_CUM_ACTUAL,
    ))
    return ctx


# ============================================================
# 第三階段：施工日報（依官方附表四格式，完全沿用既有 build_v4_report.py 邏輯）
# ============================================================

def build_daily_report(wb, meta, ctx):
    """就地執行既有的 build_v4_report.py（已完全通用、靠 meta 動態排版，
    不含任何本案專屬硬編碼），但改成函式化直接吃記憶體中的 wb/meta，
    不落地經過 json/xlsx 中繼檔。"""
    FIRST_ITEM_COL = ctx['FIRST_ITEM_COL']; N_ITEMS = ctx['N_ITEMS']
    FIRST_MAT_COL, N_MAT = ctx['FIRST_MAT_COL'], ctx['N_MAT']
    FIRST_LAB_COL, N_LABOR = ctx['FIRST_LAB_COL'], ctx['N_LABOR']
    FIRST_MCH_COL, N_MACH = ctx['FIRST_MCH_COL'], ctx['N_MACH']
    ROW_SEQ, ROW_NAME, ROW_UNIT, ROW_QTY, ROW_PRICE, ROW_LABEL, ROW_REMARK = (
        ctx['ROW_SEQ'], ctx['ROW_NAME'], ctx['ROW_UNIT'], ctx['ROW_QTY'], ctx['ROW_PRICE'], ctx['ROW_LABEL'],
        ctx['ROW_REMARK'])
    FIRST_DAY_ROW, LAST_DAY_ROW = ctx['FIRST_DAY_ROW'], ctx['LAST_DAY_ROW']
    FIRST_TXT_COL = ctx['FIRST_TXT_COL']
    TXT_FIELDS = ctx['TXT_FIELDS']
    CATEGORIES = [[label, sn, en, tp, cl] for label, sn, en, tp, cl, c1, c2 in ctx['CATEGORIES_full']]

    # 委派給既有、已充分驗證且完全通用的 build_v4_report.py 邏輯：
    # 這裡動態載入該檔案的原始碼字串，在同一個 wb 物件上執行，避免重寫一份幾百行、
    # 已經測過的排版程式碼、也避免兩份程式碼日後改一邊忘記改另一邊。
    report_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '_report_layout.py')
    if not os.path.exists(report_path):
        print('⚠️ 找不到 _report_layout.py（施工日報版面模組），略過「施工日報」分頁產生，'
              '只會輸出 日資料庫/預定進度/甘特圖 三個分頁。')
        return
    ns = dict(wb=wb, FIRST_ITEM_COL=FIRST_ITEM_COL, N_ITEMS=N_ITEMS,
              FIRST_MAT_COL=FIRST_MAT_COL, N_MAT=N_MAT, FIRST_LAB_COL=FIRST_LAB_COL, N_LABOR=N_LABOR,
              FIRST_MCH_COL=FIRST_MCH_COL, N_MACH=N_MACH, ROW_SEQ=ROW_SEQ, ROW_NAME=ROW_NAME, ROW_UNIT=ROW_UNIT,
              ROW_QTY=ROW_QTY, ROW_PRICE=ROW_PRICE, ROW_LABEL=ROW_LABEL, ROW_REMARK=ROW_REMARK,
              FIRST_DAY_ROW=FIRST_DAY_ROW, LAST_DAY_ROW=LAST_DAY_ROW, FIRST_TXT_COL=FIRST_TXT_COL,
              TXT_FIELDS=TXT_FIELDS, CATEGORIES=CATEGORIES,
              TOP_NUMERALS=TOP_NUMERALS, MID_NUMERALS=MID_NUMERALS)
    src = open(report_path, encoding='utf-8').read()
    exec(compile(src, report_path, 'exec'), ns)


# ============================================================
# 第三階段：儀表板（整體/大項中項/工項 預定vs實際進度比較，標示可能落後的項目）
# ============================================================

def build_dashboard_sheet(wb, meta, ctx):
    """新增「儀表板」分頁，目的：讓使用者一眼看出目前整體進度、以及哪些大項/中項/工項
    可能落後。所有數字都是公式，直接參照「日資料庫」與「預定進度」兩個既有分頁，
    不另外複製一份資料，之後使用者每天填報時這頁會自動跟著更新。

    ⚠️ 重要假設（寫在分頁最上方的說明區塊，這裡先記錄依據）：
    目前系統的「預定進度」分頁只讓使用者填一份『全案共用』的任務清單（起訖日+權重），
    算出的是『整個工程』逐日累計%的單一條S曲線，並沒有要求使用者替每個大項/中項/工項
    分別建立各自獨立的排程。因此本頁在判斷「哪些工項/類別可能落後」時，一律拿『全案
    整體預定進度%』當作每一個項目共同的比較基準，這只能反映『整體工程步調』，不是
    該項目自己被核定的排程——如果某工項本來就規劃在工期後段才開始，用這個基準比較
    可能會被誤判為落後，請使用者依實際狀況人工複核，不要照單全收。
    這個限制已經在分頁的說明文字中明確告知使用者，且落後門檻可由使用者自行調整（預設
    10個百分點）。若之後使用者需要「每個工項有各自獨立的預定進度」，需要另外擴充「預定
    進度」分頁的資料結構（例如把工項清單也當成可設起訖日的任務），屬於更大的功能，
    目前尚未實作。
    """
    bold = Font(bold=True)
    title_font = Font(bold=True, size=16)
    small = Font(size=10)
    note_font = Font(italic=True, color='C00000', size=9)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_top = Alignment(horizontal='left', vertical='top', wrap_text=True)
    left_mid = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='BFBFBF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill('solid', fgColor='D9E1F2')
    input_fill = PatternFill('solid', fgColor='FFF2CC')
    red_fill = PatternFill('solid', fgColor='FFC7CE'); red_font = Font(color='9C0006')
    yellow_fill = PatternFill('solid', fgColor='FFEB9C'); yellow_font = Font(color='9C6500')
    green_fill = PatternFill('solid', fgColor='C6EFCE'); green_font = Font(color='006100')

    FIRST_DAY_ROW, LAST_DAY_ROW = ctx['FIRST_DAY_ROW'], ctx['LAST_DAY_ROW']
    SCHED_SHEET_NAME = ctx['SCHED_SHEET_NAME']
    SCHED_TOTAL_COL_LETTER = ctx['SCHED_TOTAL_COL_LETTER']
    CATEGORIES_full = ctx['CATEGORIES_full']
    FIRST_ITEM_COL, N_ITEMS = ctx['FIRST_ITEM_COL'], ctx['N_ITEMS']
    ROW_NAME, ROW_UNIT, ROW_QTY = ctx['ROW_NAME'], ctx['ROW_UNIT'], ctx['ROW_QTY']
    grand_total = CATEGORIES_full[0][3]
    actual_col_letter = CATEGORIES_full[0][4]

    ws = wb.create_sheet('儀表板')
    ws.sheet_view.showGridLines = False

    QDATE = '$B$6'
    THRESH = '$E$6'

    def dmatch():
        return f"MATCH({QDATE},'日資料庫'!$B${FIRST_DAY_ROW}:$B${LAST_DAY_ROW},1)"

    def dlookup(col_letter, sheet="日資料庫"):
        return (f"IFERROR(INDEX('{sheet}'!${col_letter}${FIRST_DAY_ROW}:${col_letter}${LAST_DAY_ROW},"
                f"{dmatch()}),0)")

    def dcum(col_letter):
        return (f"IFERROR(SUM('日資料庫'!${col_letter}${FIRST_DAY_ROW}:"
                f"INDEX('日資料庫'!${col_letter}${FIRST_DAY_ROW}:${col_letter}${LAST_DAY_ROW},{dmatch()})),0)")

    # ---- Row1 標題 ----
    ws.merge_cells('A1:H1')
    ws['A1'] = f'{meta["project_name"]}　工程進度儀表板'
    ws['A1'].font = title_font

    # ---- Row2-4 假設與資料來源說明 ----
    ws.merge_cells('A2:H4')
    note = ws.cell(row=2, column=1, value=(
        '【資料來源】各大項/中項/總計的進度％，沿用「日資料庫」既有的按契約金額加權累計完成率公式；'
        '各工項的進度％改用「累計完成數量／契約數量」（數量口徑，兩者口徑雖不同，但都是拿來跟同一個'
        '整體預定進度％比較，不影響落後與否的判斷方向）。「預定進度％」則直接引用「預定進度」分頁依'
        '使用者填的任務起訖日算出的全案S曲線。\n'
        '【假設與限制】目前系統只有『全案一條』預定進度S曲線，沒有要求逐項/逐大項中項各自建立排程，'
        '所以下面每一列的「預定進度％」都是套用同一個全案整體基準，只能反映整體工程步調，不是該項目'
        '自己被核定的排程——若某工項本來就規劃在工期後段才開始，可能被誤判為落後，請人工複核，不要'
        '照單全收。「落後百分點」= 預定進度％－實際進度％，超過下方「可接受落後門檻」標記為「落後」'
        '（紅），未超過門檻但仍落後標記「留意」（黃），持平或超前標記「正常」（綠），門檻可自行調整。'
    ))
    note.font = note_font
    note.alignment = left_top
    for r in (2, 3, 4):
        ws.row_dimensions[r].height = 30

    # ---- Row6 查詢日期／門檻 輸入列 ----
    ws['A6'] = '查詢日期'; ws['A6'].font = bold; ws['A6'].alignment = left_mid
    ws.merge_cells('B6:C6')
    qd = ws['B6']; qd.value = '=TODAY()'; qd.number_format = 'yyyy/mm/dd'
    qd.fill = input_fill; qd.font = Font(bold=True); qd.alignment = center; qd.border = border
    ws['D6'] = '可接受落後門檻(百分點)'; ws['D6'].font = bold; ws['D6'].alignment = left_mid
    th = ws['E6']; th.value = 0.1; th.number_format = '0%'
    th.fill = input_fill; th.font = Font(bold=True); th.alignment = center; th.border = border
    ws.cell(row=6, column=6, value='←都可自行改成想查詢的日期／想放寬或收緊的門檻').font = note_font
    ws.merge_cells(start_row=6, start_column=6, end_row=6, end_column=8)

    # ---- Row7 狀態提示 ----
    last_day_ref = f"'日資料庫'!$B${LAST_DAY_ROW}"
    status_formula = (
        f'=IF({QDATE}<\'日資料庫\'!$B$2,'
        f'"⚠ 查詢日期早於開工日期，專案尚未開始，以下數字僅供參考。",'
        f'IF({QDATE}>{last_day_ref},'
        f'"⚠ 查詢日期超出「日資料庫」目前已建立的天數範圍（很可能是產生檔案時 --days 參數不夠大，'
        f'例如契約工期750天卻用預設200天產生，需重新以足夠的 --days 產生檔案），'
        f'以下數字僅計算到已建立範圍的最後一天為止。",""))'
    )
    ws.merge_cells('A7:H7')
    st = ws.cell(row=7, column=1, value=status_formula)
    st.font = Font(color='C00000', bold=True); st.alignment = left_mid

    # ---- Row9-10 KPI ----
    kpi_labels = ['契約總金額', '預定累計進度%(整體基準)', '實際累計進度%(全案)',
                  '落後/超前(百分點)', '尚餘工期(天)', '預定完工日']
    for j, lab in enumerate(kpi_labels):
        c = ws.cell(row=9, column=1 + j, value=lab)
        c.font = bold; c.fill = head_fill; c.alignment = center; c.border = border
    ws.cell(row=10, column=1, value=grand_total).number_format = '#,##0'
    pred_cell = ws.cell(row=10, column=2, value=f"={dlookup(SCHED_TOTAL_COL_LETTER, SCHED_SHEET_NAME)}")
    pred_cell.number_format = '0.00%'
    actual_cell = ws.cell(row=10, column=3, value=f"={dlookup(actual_col_letter)}")
    actual_cell.number_format = '0.00%'
    gap_cell = ws.cell(row=10, column=4, value='=$B$10-$C$10')
    gap_cell.number_format = '0.00%'
    ws.cell(row=10, column=5,
            value=f"=MAX('日資料庫'!$D$2-({QDATE}-'日資料庫'!$B$2+1),0)")
    ws.cell(row=10, column=6, value="='日資料庫'!$F$2").number_format = 'yyyy/mm/dd'
    for j in range(1, 7):
        c = ws.cell(row=10, column=j); c.font = Font(bold=True, size=12); c.alignment = center; c.border = border
    ws.row_dimensions[9].height = 30
    ws.row_dimensions[10].height = 22
    PRED_CELL_REF = '$B$10'  # 全案整體預定進度%，下面各列統一引用這個當基準

    cur_row = 12

    # ---- 大項/中項/總計 進度總覽 ----
    ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=7)
    ws.cell(row=cur_row, column=1,
            value='大項/中項/總計 進度總覽（中項預設已摺疊，點列號左側「+」展開）：').font = bold
    cur_row += 1

    cat_cols_block = ['類別', '名稱', '契約金額', '預定進度%(基準)', '實際進度%', '落後百分點', '狀態']

    def _cat_level(label_text):
        ch = label_text[0] if label_text else ''
        if ch in TOP_NUMERALS:
            return 1
        if ch in MID_NUMERALS:
            return 2
        return 0

    def _tag_name(label_text):
        if '、' in label_text:
            tag, name = label_text.split('、', 1)
        elif label_text.startswith('總計'):
            tag, name = '總計', '全案'
        else:
            tag, name = label_text, ''
        return tag, (name if name else label_text)

    def _write_cat_rows(cats, start_row, bold_all):
        for j, cname in enumerate(cat_cols_block):
            c = ws.cell(row=start_row, column=1 + j, value=cname)
            c.font = bold; c.fill = head_fill; c.alignment = center; c.border = border
        r = start_row + 1
        for (label, sn, en, total_price, col_letter, c1, c2) in cats:
            tag, name = _tag_name(label)
            ws.cell(row=r, column=1, value=tag)
            ws.cell(row=r, column=2, value=name)
            ws.cell(row=r, column=3, value=total_price).number_format = '#,##0'
            ws.cell(row=r, column=4, value=f'={PRED_CELL_REF}').number_format = '0.00%'
            ws.cell(row=r, column=5, value=f'={dlookup(col_letter)}').number_format = '0.00%'
            ws.cell(row=r, column=6, value=f'=D{r}-E{r}').number_format = '0.00%'
            ws.cell(row=r, column=7, value=f'=IF(F{r}>{THRESH},"落後",IF(F{r}>0,"留意","正常"))')
            for j in range(1, 8):
                c = ws.cell(row=r, column=j)
                c.border = border
                c.font = Font(bold=True) if bold_all else small
                c.alignment = left_mid if j == 2 else center
            r += 1
        return r - 1

    top_cats = [c for c in CATEGORIES_full if _cat_level(c[0]) in (0, 1)]
    mid_cats = [c for c in CATEGORIES_full if _cat_level(c[0]) == 2]

    TOP_HEAD_ROW = cur_row
    TOP_END_ROW = _write_cat_rows(top_cats, TOP_HEAD_ROW, bold_all=True)
    cat_data_ranges = [(TOP_HEAD_ROW + 1, TOP_END_ROW)]
    cur_row = TOP_END_ROW + 1

    if mid_cats:
        MID_HEAD_ROW = cur_row
        MID_END_ROW = _write_cat_rows(mid_cats, MID_HEAD_ROW, bold_all=False)
        cat_data_ranges.append((MID_HEAD_ROW + 1, MID_END_ROW))
        ws.sheet_properties.outlinePr = Outline(summaryBelow=False)
        for r in range(MID_HEAD_ROW, MID_END_ROW + 1):
            ws.row_dimensions[r].outlineLevel = 1
            ws.row_dimensions[r].hidden = True
        cur_row = MID_END_ROW + 1

    CAT_SECTION_END_ROW = cur_row - 1

    for r1, r2 in cat_data_ranges:
        rng = f'A{r1}:G{r2}'
        ws.conditional_formatting.add(rng, FormulaRule(formula=[f'$F{r1}>{THRESH}'], fill=red_fill, font=red_font))
        ws.conditional_formatting.add(rng, FormulaRule(formula=[f'$F{r1}>0'], fill=yellow_fill, font=yellow_font))
        ws.conditional_formatting.add(rng, FormulaRule(formula=[f'$F{r1}<=0'], fill=green_fill, font=green_font))

    cur_row += 1

    # ---- 工項 進度明細（全部工項，可用欄位篩選/排序找出落後工項） ----
    ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=9)
    ws.cell(row=cur_row, column=1,
            value='工項進度明細（可點欄位標題的篩選鈕依「落後百分點」排序，快速找出可能落後的工項）：').font = bold
    cur_row += 1

    item_head_row = cur_row
    item_cols_block = ['項次', '工項名稱', '單位', '契約數量', '累計完成數量',
                        '實際完成率%', '預定進度%(基準)', '落後百分點', '狀態']
    for j, cname in enumerate(item_cols_block):
        c = ws.cell(row=item_head_row, column=1 + j, value=cname)
        c.font = bold; c.fill = head_fill; c.alignment = center; c.border = border

    item_cols = [get_column_letter(FIRST_ITEM_COL + i) for i in range(N_ITEMS)]
    item_data_start = item_head_row + 1
    for idx, cl in enumerate(item_cols):
        r = item_data_start + idx
        ws.cell(row=r, column=1, value=idx + 1)
        ws.cell(row=r, column=2, value=f"='日資料庫'!{cl}{ROW_NAME}")
        ws.cell(row=r, column=3, value=f"='日資料庫'!{cl}{ROW_UNIT}")
        ws.cell(row=r, column=4, value=f"='日資料庫'!{cl}{ROW_QTY}").number_format = '#,##0.##'
        ws.cell(row=r, column=5, value=f"={dcum(cl)}").number_format = '#,##0.##'
        ws.cell(row=r, column=6, value=f'=IFERROR(E{r}/D{r},0)').number_format = '0.00%'
        ws.cell(row=r, column=7, value=f'={PRED_CELL_REF}').number_format = '0.00%'
        ws.cell(row=r, column=8, value=f'=G{r}-F{r}').number_format = '0.00%'
        ws.cell(row=r, column=9, value=f'=IF(H{r}>{THRESH},"落後",IF(H{r}>0,"留意","正常"))')
        for j in range(1, 10):
            c = ws.cell(row=r, column=j)
            c.font = small
            c.alignment = left_mid if j == 2 else center
    item_data_end = item_data_start + N_ITEMS - 1

    item_rng = f'A{item_head_row}:I{item_data_end}'
    tbl = Table(displayName='工項進度明細表', ref=item_rng)
    tbl.tableStyleInfo = TableStyleInfo(name='TableStyleMedium2', showRowStripes=True)
    ws.add_table(tbl)

    body_rng = f'A{item_data_start}:I{item_data_end}'
    ws.conditional_formatting.add(body_rng, FormulaRule(formula=[f'$H{item_data_start}>{THRESH}'],
                                                          fill=red_fill, font=red_font))
    ws.conditional_formatting.add(body_rng, FormulaRule(formula=[f'$H{item_data_start}>0'],
                                                          fill=yellow_fill, font=yellow_font))
    ws.conditional_formatting.add(body_rng, FormulaRule(formula=[f'$H{item_data_start}<=0'],
                                                          fill=green_fill, font=green_font))

    ws.freeze_panes = ws.cell(row=item_data_start, column=1)
    # 欄寬需同時兼顧上面「KPI／大項中項總覽」（放得下千分位金額、較長的欄位標題）跟下面
    # 「工項明細表」（項次/單位這種短欄位）兩種不同用途，取兩邊都能正常顯示的折衷寬度。
    widths = {1: 17, 2: 26, 3: 15, 4: 14, 5: 14, 6: 13, 7: 13, 8: 12, 9: 10}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.sheet_view.showOutlineSymbols = True
    # 預設列印範圍只框「總覽＋大項/中項摘要」這一小段（跟日資料庫/預定進度分頁的作法一致），
    # 不含下面完整的工項明細表——工項明細表是給Excel篩選/排序在螢幕上用的，筆數可能有上百列，
    # 硬塞進列印頁會被壓縮到看不清楚；如果需要印工項明細，請自行在Excel框選範圍後另外列印。
    ws.print_area = f'A1:G{CAT_SECTION_END_ROW}'
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1


# ============================================================
# 第三階段：PCCES來源核對表（供debug/人工比對用）
# ============================================================

def build_source_ref_sheet(wb, meta, source_path, used_sheet, records, items, CATEGORIES, N_ITEMS,
                            grand_total, grand_total_source, self_sum):
    """把程式從PCCES檔案實際解析出來的每一列原始資料（含大項/中項/小計/總計標題列，
    不只是80/N個工項本身）攤開列在一張新分頁裡，方便使用者拿這張表跟自己手上的
    PCCES螢幕畫面或列印稿逐列核對，快速看出是不是選錯分頁、漏抓欄位、或大項/中項
    編號沒偵測到——不需要每次都回來這個對話讓我重新讀一次原始檔案才能debug。"""
    ws = wb.create_sheet('PCCES來源核對(debug用)')
    thin = Side(style='thin', color='BFBFBF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill('solid', fgColor='D9E1F2')
    warn_fill = PatternFill('solid', fgColor='FFC7CE')
    bold = Font(name=FONT_NAME, bold=True)
    normal = Font(name=FONT_NAME)

    r = 1
    ws.cell(row=r, column=1, value='PCCES來源資料核對表（本系統擴充，非官方施工日誌固定分頁，純供人工debug比對用）').font = Font(name=FONT_NAME, bold=True, size=13)
    r += 2
    info_rows = [
        ('來源檔案', source_path),
        ('實際使用分頁', used_sheet),
        ('抓到的工程名稱', meta['project_name']),
        ('抓到的施工地點', meta['location']),
        ('抓到的工程編號', meta['proj_code']),
        ('抓到的主辦機關', meta['agency']),
        ('工項總數 N_ITEMS', N_ITEMS),
    ]
    for label, val in info_rows:
        ws.cell(row=r, column=1, value=label).font = bold
        ws.cell(row=r, column=2, value=val).font = normal
        r += 1
    r += 1

    ws.cell(row=r, column=1, value='契約總價來源判斷').font = bold
    r += 1
    rows2 = [
        ('全部工項複價自行加總', self_sum),
        ('目前採用的契約總價', grand_total),
        ('採用來源說明', grand_total_source),
        ('差額（自行加總－採用值）', round(self_sum - grand_total, 2)),
    ]
    mismatch = abs(self_sum - grand_total) > 1
    for label, val in rows2:
        c1 = ws.cell(row=r, column=1, value=label)
        c2 = ws.cell(row=r, column=2, value=val)
        c1.font = bold
        c2.font = normal
        if isinstance(val, (int, float)):
            c2.number_format = '#,##0'
        if mismatch and label.startswith('差額'):
            c1.fill = warn_fill
            c2.fill = warn_fill
        r += 1
    if mismatch:
        ws.cell(row=r, column=1,
                value='⚠️ 上面差額不是0：常見原因是PCCES文件裡「總價」列之後還有另計的扣款/調整'
                      '項次（例如剩餘價值折價費），依業務慣例不算入契約金額——請對照下方「大項/中項'
                      '偵測結果」與原始工項清單，確認是不是這種情況；如果不是，可能是「總價」列本身'
                      '跳號或非標準寫法，也請一併核對。').font = Font(name=FONT_NAME, color='C00000')
        ws.row_dimensions[r].height = 30
        r += 1
    r += 1

    ws.cell(row=r, column=1, value='大項/中項偵測結果（自算金額＝該範圍內所有工項複價加總）').font = bold
    r += 1
    hdr_r = r
    for c, label in enumerate(['類別', '起始工項#', '結束工項#', '自算金額'], start=1):
        cell = ws.cell(row=hdr_r, column=c, value=label)
        cell.font = bold
        cell.fill = head_fill
        cell.border = border
    r += 1
    for label, sn, en, tp, *_rest in CATEGORIES:
        for c, val in enumerate([label, sn, en, tp], start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = border
            cell.font = normal
            if c == 4 and isinstance(val, (int, float)):
                cell.number_format = '#,##0'
        r += 1
    r += 1

    ws.cell(row=r, column=1,
            value='原始工項清單（含大項/中項/小計/總計標題列，跟PCCES檔案逐列對照用）').font = bold
    r += 1
    item_hdr_r = r
    headers = ['第幾列(原始順序)', '項次', '項目及說明', '單位', '數量', '單價', '複價', '編碼備註', '判斷類型', '是否為大項/中項番號']
    for c, label in enumerate(headers, start=1):
        cell = ws.cell(row=item_hdr_r, column=c, value=label)
        cell.font = bold
        cell.fill = head_fill
        cell.border = border
    r += 1
    seq = 0
    for rec in records:
        seq += 1
        numeral = str(rec['項次']).strip() if not _is_blank(rec['項次']) else ''
        is_marker = '大項' if numeral in TOP_NUMERALS else ('中項' if numeral in MID_NUMERALS else '')
        vals = [seq, rec['項次'], rec['項目及說明'], rec['單位'], rec['數量'], rec['單價'],
                rec['複價'], rec['編碼備註'], rec['類型'], is_marker]
        for c, val in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = border
            cell.font = normal
            if c in (5, 6, 7) and isinstance(val, (int, float)):
                cell.number_format = '#,##0.##'
            if is_marker:
                cell.fill = PatternFill('solid', fgColor='FFF2CC')
        r += 1

    widths = {1: 24, 2: 14, 3: 34, 4: 16, 5: 10, 6: 10, 7: 12, 8: 16, 9: 10, 10: 12}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = f'A{item_hdr_r + 1}'
    ws.sheet_view.showGridLines = False
    # 這張是debug用參考表，主要用途是在Excel裡開啟後左右/上下捲動比對，不是要印出來的
    # 正式表單；仍設定橫向+fitToWidth=1，避免萬一列印時被縱向欄寬硬切成一堆只有兩三欄的
    # 內容不連貫的分頁。
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True


# ============================================================
# CLI
# ============================================================

# ============================================================
# 第四階段：工期展延(輸入展延後完工日→自動反推總天數) + 契約變更快速套用
#   （把「變更後」的PCCES檔案套進一份已經填了逐日資料的既有施工日誌，不整份重建、
#    不遺失已填資料）
# ============================================================

def resolve_days(start_date, days_arg=None, end_date_arg=None, min_days=None):
    """算出日資料庫要建幾天份。優先順序：
      1. 有給 end_date_arg（展延/預定完工日）→ 天數 = 開工日~此日期(含頭尾)。
      2. 否則用 days_arg（沒給就預設200）。
      3. 若算出來的天數比 min_days 少（通常是「舊檔案已經填到第幾天」），
         為了不讓已填資料被截斷遺失，自動延長到 min_days，並在來源說明文字裡註明。
    回傳 (days, end_date, source_text)。"""
    if end_date_arg:
        end_date = (datetime.datetime.strptime(end_date_arg, '%Y-%m-%d').date()
                    if isinstance(end_date_arg, str) else end_date_arg)
        days = (end_date - start_date).days + 1
        if days < 1:
            raise ValueError(f'展延後完工日期({end_date})比開工日期({start_date})還早，'
                              f'算出來的工期天數是負的，請確認日期有沒有填反。')
        source = f'開工日{start_date} ~ 展延後完工日{end_date}，共{days}天（=展延後完工日-開工日+1）'
    else:
        days = int(days_arg) if days_arg else 200
        end_date = start_date + datetime.timedelta(days=days - 1)
        source = f'--days參數指定{days}天（未提供展延後完工日期）'
    if min_days and days < min_days:
        days = min_days
        end_date = start_date + datetime.timedelta(days=days - 1)
        source += f'；但比舊檔案已經填到的第{min_days}天還少，為了不遺失已填資料，已自動延長為{days}天'
    return days, end_date, source


def _find_label_run_columns(ws, row, label_text, min_col=1, max_col=600):
    """在指定列(row)裡找出所有值等於label_text的欄，並把「彼此緊鄰(欄號連續)」的
    分成一段一段回傳 [(start_col, end_col), ...]（依左到右順序）。用途：日資料庫的
    工項/材料/工別/機具欄位區塊在產生時一定是連續整段寫入同一個ROW_LABEL文字，
    靠這個特徵就能不依賴任何欄位數常數、直接在既有檔案裡把區塊位置找回來——即使
    材料跟機具用的是同一句「本日使用數量」文字也不會混在一起，因為中間隔著工別
    欄位(文字是「本日人數」)，欄號不連續，會被切成兩段。"""
    cols = [c for c in range(min_col, max_col + 1) if ws.cell(row=row, column=c).value == label_text]
    if not cols:
        return []
    runs, run_start, prev = [], cols[0], cols[0]
    for c in cols[1:]:
        if c == prev + 1:
            prev = c
        else:
            runs.append((run_start, prev))
            run_start, prev = c, c
    runs.append((run_start, prev))
    return runs


def extract_daily_records(old_path):
    """讀取一份「本工具產生的施工日誌.xlsx」（可能已經填了很多天的逐日資料），把使用者
    手動填寫過的所有資料原樣抽取出來，供 apply_change_order() 套用契約變更時搬回新檔案。

    刻意設計成完全靠「儲存格內容特徵」（欄位標籤文字、欄與欄的相對順序、固定不變的
    說明文字）在活頁簿裡自行定位，不依賴任何外部 _meta.json 側車檔案——這樣即使使用者
    手上只有單一份.xlsx檔案（meta.json 弄丟了）也能正常運作，更穩固。

    已知限制：如果使用者曾經手動大幅更動過「日資料庫」分頁的版面（例如自行插入/搬移欄、
    改掉欄位標籤文字），比對可能會失敗或找錯欄位，此時會丟出例外訊息說明無法辨識，
    請使用者不要手動改動版面結構（填資料本身完全不受影響）。
    """
    wb = openpyxl.load_workbook(old_path, data_only=False)
    if '日資料庫' not in wb.sheetnames:
        raise ValueError('這份檔案裡找不到「日資料庫」分頁，看起來不是本工具產生的施工日誌，'
                          '無法套用契約變更（請確認上傳的是「既有的施工日誌.xlsx」，而不是PCCES原始檔）。')
    ws = wb['日資料庫']
    ROW_REMARK, ROW_SEQ, ROW_NAME, ROW_UNIT, ROW_QTY, ROW_PRICE, ROW_LABEL = 3, 4, 5, 6, 7, 8, 9
    FIRST_DAY_ROW = 10

    old_meta = {
        'project_name': ws['B1'].value, 'proj_code': ws['G1'].value,
        'grand_total': ws['I1'].value, 'agency': ws['K1'].value,
        'location': ws['H2'].value,
    }
    start_date_val = ws['B2'].value
    if isinstance(start_date_val, datetime.datetime):
        start_date_val = start_date_val.date()
    old_meta['start_date'] = start_date_val

    # 掃描A欄「第幾天」序號實際填到第幾列，比D2寫的「工期(日曆天)」更可靠——使用者可能
    # 超過原訂工期還在繼續填、或D2本身被手動改過而跟實際列數對不上。
    max_day_row = FIRST_DAY_ROW - 1
    r = FIRST_DAY_ROW
    while ws.cell(row=r, column=1).value not in (None, ''):
        max_day_row = r
        r += 1
    old_days = max_day_row - FIRST_DAY_ROW + 1
    LAST_DAY_ROW = max_day_row if old_days > 0 else FIRST_DAY_ROW - 1

    def read_daily(col):
        out = {}
        for rr in range(FIRST_DAY_ROW, LAST_DAY_ROW + 1):
            v = ws.cell(row=rr, column=col).value
            if v not in (None, ''):
                out[rr - FIRST_DAY_ROW] = v
        return out

    item_runs = _find_label_run_columns(ws, ROW_LABEL, '本日完成數量')
    if not item_runs:
        raise ValueError('在「日資料庫」分頁裡找不到「本日完成數量」欄位標籤，欄位結構跟預期不符，'
                          '無法自動比對工項，請確認這份檔案的版面沒有被手動改動過。')
    item_start, item_end = item_runs[0]
    items_old = []
    for col in range(item_start, item_end + 1):
        items_old.append({
            'name': ws.cell(row=ROW_NAME, column=col).value,
            'unit': ws.cell(row=ROW_UNIT, column=col).value,
            'qty': ws.cell(row=ROW_QTY, column=col).value,
            'price': ws.cell(row=ROW_PRICE, column=col).value,
            'daily': read_daily(col),
        })

    usage_runs = _find_label_run_columns(ws, ROW_LABEL, '本日使用數量', min_col=item_end + 1)
    labor_runs = _find_label_run_columns(ws, ROW_LABEL, '本日人數', min_col=item_end + 1)
    mat_range = usage_runs[0] if len(usage_runs) >= 1 else None
    mach_range = usage_runs[1] if len(usage_runs) >= 2 else None
    labor_range = labor_runs[0] if labor_runs else None

    def read_block(rng):
        if rng is None:
            return []
        s, e = rng
        return [{'name': ws.cell(row=ROW_NAME, column=col).value,
                  'unit': ws.cell(row=ROW_UNIT, column=col).value,
                  'daily': read_daily(col)} for col in range(s, e + 1)]

    mat_old = read_block(mat_range)
    labor_old = read_block(labor_range)
    mach_old = read_block(mach_range)

    # 八大項文字欄：用每欄固定不變的「完整說明文字」(寫在 ROW_NAME 那列)精準比對定位，
    # 不受材料/工別/機具實際欄位數是否跟目前程式版本一致影響。
    txt_daily = {}
    for key, full_text, kind in TXT_FIELDS:
        found_col = None
        for col in range(item_end + 1, ws.max_column + 1):
            if ws.cell(row=ROW_NAME, column=col).value == full_text:
                found_col = col
                break
        txt_daily[key] = read_daily(found_col) if found_col else {}

    sched_tasks = []
    if '預定進度' in wb.sheetnames:
        ws2 = wb['預定進度']
        S_NAME, S_AMOUNT, S_START, S_END = 4, 5, 7, 8
        S_FIRST_COL, N_SCHED = 3, 30
        for i in range(N_SCHED):
            col = S_FIRST_COL + i
            sched_tasks.append({
                'name': ws2.cell(row=S_NAME, column=col).value,
                'amount': ws2.cell(row=S_AMOUNT, column=col).value,
                'start': ws2.cell(row=S_START, column=col).value,
                'end': ws2.cell(row=S_END, column=col).value,
            })

    dashboard_inputs = {}
    if '儀表板' in wb.sheetnames:
        wsd = wb['儀表板']
        try:
            dashboard_inputs['query_date'] = wsd['B6'].value
            dashboard_inputs['threshold'] = wsd['E6'].value
        except Exception:
            pass

    return {
        'old_meta': old_meta, 'old_days': old_days, 'items_old': items_old,
        'mat_old': mat_old, 'labor_old': labor_old, 'mach_old': mach_old,
        'txt_daily': txt_daily, 'sched_tasks': sched_tasks,
        'dashboard_inputs': dashboard_inputs,
    }


def _match_items_by_name(old_list, new_name_to_col):
    matched, unmatched = [], []
    for rec in old_list:
        name = (rec['name'] or '').strip()
        col = new_name_to_col.get(name)
        (matched if col is not None else unmatched).append((rec, col) if col is not None else rec)
    return matched, unmatched


def _build_deleted_items_sheet(wb, unmatched_old_items):
    ws = wb.create_sheet('已刪除工項歷史記錄')
    ws.sheet_view.showGridLines = False
    bold = Font(bold=True)
    thin = Side(style='thin', color='BFBFBF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill('solid', fgColor='FCE4D6')

    ws['A1'] = '已刪除工項歷史記錄'
    ws['A1'].font = Font(bold=True, size=13)
    note = ws.cell(row=2, column=1,
        value='※這些工項在最新一次契約變更後，已經不在新的PCCES工項清單裡（可能真的被刪除，也可能'
              '只是改了名稱用字，但本工具是用「工項名稱逐字完全相同」判斷是否為同一項，改名會被當成'
              '「舊項目刪除+新項目新增」——如果是這種情況，請人工核對後把數量手動搬到新名稱的項目）。'
              '下表保留這些工項在契約變更前的累計完成數量與最後填報是第幾天，僅供歷史查核使用，'
              '不再參與「日資料庫」的進度公式計算。原始的逐日筆記錄仍完整保留在您上傳的舊檔案裡——'
              '本次套用契約變更「不會修改、也不會刪除」您上傳的舊檔案本身，有需要可以隨時回頭查閱。')
    note.font = Font(italic=True, color='C00000')
    note.alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells('A2:G2')
    ws.row_dimensions[2].height = 75

    headers = ['工項名稱', '單位(舊)', '契約數量(舊)', '單價(舊)', '複價(舊)', '累計完成數量(變更前)', '最後填報第幾天']
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=j, value=h)
        c.font = bold; c.fill = head_fill; c.border = border
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for i, rec in enumerate(unmatched_old_items):
        r = 5 + i
        total_done = sum(v for v in rec['daily'].values() if isinstance(v, (int, float)))
        last_day = (max(rec['daily'].keys()) + 1) if rec['daily'] else ''
        qty, price = rec['qty'], rec['price']
        complex_price = qty * price if isinstance(qty, (int, float)) and isinstance(price, (int, float)) else ''
        vals = [rec['name'], rec['unit'], qty, price, complex_price, total_done, last_day]
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=j, value=v)
            c.border = border
            if j in (3, 4, 5, 6):
                c.number_format = '#,##0.##'

    widths = {1: 30, 2: 10, 3: 12, 4: 12, 5: 14, 6: 16, 7: 12}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = 'A5'


def apply_change_order(old_xlsx_path, new_pcces_path, sheet_name=None,
                        agency_override=None, project_name_override=None,
                        location_override=None, proj_code_override=None,
                        end_date=None, days_override=None,
                        n_mat=None, n_labor=None, n_mach=None):
    """把一份「變更後的PCCES匯出檔」套用到一份既有、已經填了逐日資料的施工日誌上，
    輸出一份全新的活頁簿——不整份重來、也絕不修改/覆寫您上傳的舊檔案本身。

    設計原則（假設，已寫進回傳的extra['warnings']與GUI畫面文字，供使用者核對）：
    - 工項清單/大項中項/契約金額全部改用這次的新PCCES檔案重新解析出來的版本，視為
      最新、最權威的來源（等於是拿新契約書重新產生一次，只是把舊的逐日填報資料原樣
      搬回來，不用整份重打）。
    - 開工日期固定沿用舊檔案（契約變更不會讓已經開工的日期往回改），不接受覆寫。
    - 已經填寫過的資料——工項每日完成數量／材料／工別／機具／日報八大項文字欄／
      預定進度任務清單——全部原樣搬回新活頁簿對應欄位。工項用「名稱逐字完全相同」
      比對是否為同一項；材料/工別/機具三類本來就跟契約工項無關（是資源主檔，不是
      契約清單），用欄位原始順序位置比對搬回。
    - 找不到對應新工項的舊工項（判斷為「已刪除」，也可能只是改名但無法自動辨識）：
      不會被丟棄，改整理進新增的「已刪除工項歷史記錄」分頁保留查核，不再參與進度計算。
    - 天數：預設沿用舊檔案「實際已經填到第幾天」；若有指定展延後完工日期或天數，
      以指定值為準，但絕不會小於舊檔案已填天數（避免截斷遺失已填資料，會自動延長
      並在警告訊息說明）。

    回傳 (wb, meta, ctx, extra)，extra 內含：
      unmatched_old_items（判定為已刪除的舊工項名稱清單）、
      unmatched_new_items（這次新增的工項名稱清單）、
      changed_items（名稱相同但單位不同、建議人工複核的工項清單）、
      warnings（其他警告文字）、old_days、new_days。
    """
    old = extract_daily_records(old_xlsx_path)

    # 自動判斷新檔案是「標準PCCES詳細價目表」還是「變更設計比較表」(壹/一/1/(1)
    # 四層結構、同列並列原契約與變更後兩組數量單價)——兩種格式的解析器回傳完全
    # 相同的tuple結構，所以下面只需要在呼叫哪個解析函式上分岔，其餘邏輯不用重複。
    if looks_like_change_comparison_file(new_pcces_path, sheet_name):
        parse_fn = parse_change_comparison_file
        print('偵測到「變更設計比較表」格式(同一列並列原契約/變更後兩組數量單價)，'
              '改用專屬解析器讀取，並優先採用議價後正式單價（如果檔案裡有議價明細表）。')
    else:
        parse_fn = parse_pcces_file

    (meta, items, CATEGORIES, N_ITEMS, grand_total, records, grand_total_source, self_sum,
     used_sheet) = parse_fn(
        new_pcces_path, sheet_name=sheet_name, agency_override=agency_override,
        project_name_override=project_name_override, location_override=location_override,
        proj_code_override=proj_code_override)

    start_date = old['old_meta']['start_date'] or datetime.date.today()
    if isinstance(start_date, datetime.datetime):
        start_date = start_date.date()

    # 套用契約變更時，若使用者沒有指定天數也沒有指定展延後完工日期，預設應該是「完全沿用舊
    # 檔案原本的日資料庫天數」，而不是退回產生全新日誌時的通用預設值(200天)——否則會在使用者
    # 沒要求的情況下把日曆表無端放大到200天，跟「套用變更」這個操作的直覺(應該盡量維持原狀，
    # 只有工項清單真的變了)不符。所以這裡把 old_days 當成「沒指定時的預設值」餵給
    # resolve_days()，而不是讓它用回自己內建的200天預設；min_days 仍然保留，用來防呆
    # ——萬一使用者自己指定了一個比舊檔案已建天數還小的天數，還是會被攔下來自動延長。
    if not days_override and not end_date:
        days = old['old_days'] or 200
        days_source = f'未指定天數/展延日期，沿用舊檔案原本的天數({days}天)不變'
    else:
        days, _end_date, days_source = resolve_days(
            start_date, days_arg=days_override, end_date_arg=end_date, min_days=old['old_days'])
    warnings = [f'工期天數依據：{days_source}']

    n_mat = n_mat if n_mat is not None else (len(old['mat_old']) or 12)
    n_labor = n_labor if n_labor is not None else (len(old['labor_old']) or 8)
    n_mach = n_mach if n_mach is not None else (len(old['mach_old']) or 8)

    wb = openpyxl.Workbook()
    ctx = build_day_database(wb, meta, items, CATEGORIES, N_ITEMS, grand_total,
                              start_date, days, n_mat, n_labor, n_mach)
    ctx = build_sched_and_gantt(wb, meta, ctx)
    build_daily_report(wb, meta, ctx)
    build_dashboard_sheet(wb, meta, ctx)
    build_source_ref_sheet(wb, meta, new_pcces_path, used_sheet, records, items, CATEGORIES,
                            N_ITEMS, grand_total, grand_total_source, self_sum)

    ws = wb['日資料庫']
    FIRST_DAY_ROW, LAST_DAY_ROW = ctx['FIRST_DAY_ROW'], ctx['LAST_DAY_ROW']
    new_name_to_col = {}
    new_name_counts = {}
    for col in range(ctx['FIRST_ITEM_COL'], ctx['LAST_ITEM_COL'] + 1):
        name = (ws.cell(row=5, column=col).value or '').strip()
        new_name_to_col.setdefault(name, col)
        new_name_counts[name] = new_name_counts.get(name, 0) + 1

    # 「工項名稱逐字完全相同」比對法本身有個已知風險：如果同一份契約清單裡剛好有
    # 兩個以上工項用了完全相同的名稱(例如同一個工程有兩處不同的施工範圍，各自都有
    # 一項「構造物開挖，砂土礫石，深度＜5m，機械挖」)，比對表(name → 欄位)只能記住
    # 其中一個欄位，會讓其中一項的逐日資料被錯誤搬到另一項欄位、另一項則完全收不到
    # 資料(顯示成「新增工項」)。這不是新解析器特有的問題(標準PCCES格式如果也有
    # 重複名稱一樣會發生)，但務必明確示警，不能讓使用者誤以為資料一定有正確對齊。
    old_name_counts = {}
    for rec in old['items_old']:
        nm = (rec['name'] or '').strip()
        old_name_counts[nm] = old_name_counts.get(nm, 0) + 1
    dup_names = sorted({nm for nm, cnt in old_name_counts.items() if cnt > 1} |
                        {nm for nm, cnt in new_name_counts.items() if cnt > 1})
    dup_names = [nm for nm in dup_names if nm]
    if dup_names:
        preview = '、'.join(f'「{nm}」' for nm in dup_names[:5])
        more = f'...等共{len(dup_names)}個' if len(dup_names) > 5 else ''
        duplicate_name_warning = (
            f'⚠️ 重要：偵測到工項名稱重複(舊檔案和/或新清單裡，同一個名稱出現超過一次)：{preview}{more}。'
            f'本工具用「名稱逐字完全相同」比對新舊工項，遇到重複名稱時只能對到其中一個，'
            f'可能會把逐日資料搬到錯的工項、或讓其中一項誤判成「新增工項」而遺失原本累計數量。'
            f'請針對這些名稱重複的工項，人工核對「日資料庫」分頁的逐日數量是否正確對應到您預期的'
            f'施工範圍/群組，必要時手動搬移數字。')
        print(duplicate_name_warning)
        warnings.append(duplicate_name_warning)

    matched, unmatched_old_items = _match_items_by_name(old['items_old'], new_name_to_col)
    matched_new_cols = set()
    changed_items = []
    for rec, col in matched:
        matched_new_cols.add(col)
        new_unit = ws.cell(row=6, column=col).value
        if rec['unit'] and new_unit and str(rec['unit']).strip() != str(new_unit).strip():
            changed_items.append({'name': rec['name'], 'old_unit': rec['unit'], 'new_unit': new_unit})
        for offset, val in rec['daily'].items():
            r = FIRST_DAY_ROW + offset
            if r <= LAST_DAY_ROW:
                c = ws.cell(row=r, column=col, value=val)
                c.number_format = '#,##0.##'

    unmatched_new_items = [ws.cell(row=5, column=col).value
                            for col in range(ctx['FIRST_ITEM_COL'], ctx['LAST_ITEM_COL'] + 1)
                            if col not in matched_new_cols]

    def restore_block(old_block, first_col, n):
        for i, rec in enumerate(old_block[:n]):
            col = first_col + i
            if rec.get('name'):
                ws.cell(row=5, column=col, value=rec['name'])
            if rec.get('unit'):
                ws.cell(row=6, column=col, value=rec['unit'])
            for offset, val in rec['daily'].items():
                r = FIRST_DAY_ROW + offset
                if r <= LAST_DAY_ROW:
                    ws.cell(row=r, column=col, value=val)

    restore_block(old['mat_old'], ctx['FIRST_MAT_COL'], ctx['N_MAT'])
    restore_block(old['labor_old'], ctx['FIRST_LAB_COL'], ctx['N_LABOR'])
    restore_block(old['mach_old'], ctx['FIRST_MCH_COL'], ctx['N_MACH'])
    if len(old['mat_old']) > ctx['N_MAT'] or len(old['labor_old']) > ctx['N_LABOR'] or len(old['mach_old']) > ctx['N_MACH']:
        warnings.append('⚠️ 舊檔案的材料/工別/機具欄位數比這次設定的上限多，超過的部分已被截斷，'
                         '如果有需要請在進階設定調高對應的筆數上限後重新套用。')

    for key, full_text, kind in ctx['TXT_FIELDS']:
        idx = [i for i, (k, _, _) in enumerate(ctx['TXT_FIELDS']) if k == key][0]
        col = ctx['FIRST_TXT_COL'] + idx
        for offset, val in old['txt_daily'].get(key, {}).items():
            r = FIRST_DAY_ROW + offset
            if r <= LAST_DAY_ROW:
                ws.cell(row=r, column=col, value=val)

    if old['sched_tasks'] and '預定進度' in wb.sheetnames:
        ws2 = wb['預定進度']
        S_NAME, S_AMOUNT, S_START, S_END = 4, 5, 7, 8
        for i, t in enumerate(old['sched_tasks']):
            col = ctx['SCHED_TASK_FIRST_COL'] + i
            if t.get('name'):
                ws2.cell(row=S_NAME, column=col, value=t['name'])
            if t.get('amount') not in (None, ''):
                c = ws2.cell(row=S_AMOUNT, column=col, value=t['amount']); c.number_format = '#,##0'
            if t.get('start') not in (None, ''):
                c = ws2.cell(row=S_START, column=col, value=t['start']); c.number_format = 'yyyy/mm/dd'
            if t.get('end') not in (None, ''):
                c = ws2.cell(row=S_END, column=col, value=t['end']); c.number_format = 'yyyy/mm/dd'

    if old['dashboard_inputs'] and '儀表板' in wb.sheetnames:
        wsd = wb['儀表板']
        try:
            if old['dashboard_inputs'].get('query_date') not in (None, ''):
                wsd['B6'] = old['dashboard_inputs']['query_date']
            if old['dashboard_inputs'].get('threshold') not in (None, ''):
                wsd['E6'] = old['dashboard_inputs']['threshold']
        except Exception:
            pass

    if unmatched_old_items:
        _build_deleted_items_sheet(wb, unmatched_old_items)

    extra = {
        'unmatched_old_items': [r['name'] for r in unmatched_old_items],
        'unmatched_new_items': unmatched_new_items,
        'changed_items': changed_items,
        'warnings': warnings,
        'old_project_name': old['old_meta'].get('project_name'),
        'old_days': old['old_days'],
        'new_days': days,
    }
    return wb, meta, ctx, extra


def main():
    ap = argparse.ArgumentParser(description='公共工程施工日誌產生器（通用版）——見檔案開頭說明文字。',
                                  formatter_class=argparse.RawDescriptionHelpFormatter,
                                  epilog=__doc__)
    ap.add_argument('pcces_file', help='PCCES匯出的詳細價目表/標價清單/預算詳細表（.xls 或 .xlsx）；'
                                        '若有給 --apply-change-to，這份檔案代表「變更後」的新工項清單')
    ap.add_argument('--sheet', default=None, help='手動指定分頁名稱（預設自動偵測）')
    ap.add_argument('--start-date', default=None, help='開工日期 YYYY-MM-DD（預設今天；套用契約變更時此參數會被忽略，固定沿用舊檔案的開工日）')
    ap.add_argument('--days', type=int, default=None, help='日資料庫要建幾天份（預設200；有給 --end-date 時此參數被忽略）')
    ap.add_argument('--end-date', default=None, help='展延後完工日期 YYYY-MM-DD（如果提供，優先於--days，天數改用「開工日~此日期」自動反推，並在畫面印出算式）')
    ap.add_argument('--agency', default=None, help='主辦機關（PCCES讀不到時手動補）')
    ap.add_argument('--project-name', default=None, help='工程名稱（手動覆寫）')
    ap.add_argument('--location', default=None, help='施工地點（手動覆寫）')
    ap.add_argument('--proj-code', default=None, help='工程編號（手動覆寫）')
    ap.add_argument('--n-mat', type=int, default=12, help='材料主檔筆數上限（預設12）')
    ap.add_argument('--n-labor', type=int, default=8, help='工別主檔筆數上限（預設8）')
    ap.add_argument('--n-mach', type=int, default=8, help='機具主檔筆數上限（預設8）')
    ap.add_argument('--out', default=None, help='輸出檔名（預設用工程名稱自動命名）')
    ap.add_argument('--apply-change-to', default=None, metavar='舊施工日誌.xlsx',
                     help='套用契約變更：填入既有、已經填了逐日資料的施工日誌.xlsx路徑，這次的pcces_file'
                          '會被當成「變更後」的新契約清單套用進去，輸出全新檔案、不會動到這份舊檔案本身')
    args = ap.parse_args()

    if args.apply_change_to:
        wb, meta, ctx, extra = apply_change_order(
            args.apply_change_to, args.pcces_file, sheet_name=args.sheet,
            agency_override=args.agency, project_name_override=args.project_name,
            location_override=args.location, proj_code_override=args.proj_code,
            end_date=args.end_date, days_override=args.days,
            n_mat=args.n_mat, n_labor=args.n_labor, n_mach=args.n_mach)
        safe_name = re.sub(r'[\\/:*?"<>|]', '_', meta['project_name'] or 'CONSTRUCTION_LOG')
        out_path = args.out or f'{safe_name}_施工日誌(契約變更).xlsx'
        wb.save(out_path)
        print(f'\n完成！契約變更已套用，已輸出全新檔案：{out_path}（原始舊檔案 {args.apply_change_to} 未被修改）')
        for w in extra['warnings']:
            print(f'ℹ️ {w}')
        if extra['unmatched_new_items']:
            print(f"➕ 這次新增的工項（{len(extra['unmatched_new_items'])}項）：" + '、'.join(extra['unmatched_new_items'][:10])
                  + ('...' if len(extra['unmatched_new_items']) > 10 else ''))
        if extra['unmatched_old_items']:
            print(f"➖ 判定為已刪除的舊工項（{len(extra['unmatched_old_items'])}項，已移到「已刪除工項歷史記錄」分頁）：" +
                  '、'.join(extra['unmatched_old_items'][:10]) + ('...' if len(extra['unmatched_old_items']) > 10 else ''))
        if extra['changed_items']:
            print(f"⚠️ 名稱相同但單位不同、建議人工複核（{len(extra['changed_items'])}項）：" +
                  '、'.join(f"{c['name']}({c['old_unit']}→{c['new_unit']})" for c in extra['changed_items'][:10]))
        return

    (meta, items, CATEGORIES, N_ITEMS, grand_total, records, grand_total_source, self_sum,
     used_sheet) = parse_pcces_file(
        args.pcces_file, sheet_name=args.sheet, agency_override=args.agency,
        project_name_override=args.project_name, location_override=args.location,
        proj_code_override=args.proj_code)

    start_date = (datetime.datetime.strptime(args.start_date, '%Y-%m-%d').date()
                  if args.start_date else datetime.date.today())
    days, _end_date, days_source = resolve_days(start_date, days_arg=args.days, end_date_arg=args.end_date)
    print(f'ℹ️ 工期天數依據：{days_source}')

    wb = openpyxl.Workbook()
    ctx = build_day_database(wb, meta, items, CATEGORIES, N_ITEMS, grand_total,
                              start_date, days, args.n_mat, args.n_labor, args.n_mach)
    ctx = build_sched_and_gantt(wb, meta, ctx)
    build_daily_report(wb, meta, ctx)
    build_dashboard_sheet(wb, meta, ctx)
    build_source_ref_sheet(wb, meta, args.pcces_file, used_sheet, records, items, CATEGORIES,
                            N_ITEMS, grand_total, grand_total_source, self_sum)

    safe_name = re.sub(r'[\\/:*?"<>|]', '_', meta['project_name'] or 'CONSTRUCTION_LOG')
    out_path = args.out or f'{safe_name}_施工日誌.xlsx'
    wb.save(out_path)

    meta_out = dict(ctx)
    meta_out['CATEGORIES'] = [[label, sn, en, tp, cl] for label, sn, en, tp, cl, c1, c2 in ctx['CATEGORIES_full']]
    del meta_out['CATEGORIES_full']
    meta_out.update(project_name=meta['project_name'], location=meta['location'],
                     proj_code=meta['proj_code'], agency=meta['agency'], grand_total=grand_total)
    meta_out['start_date'] = str(start_date)
    meta_path = os.path.splitext(out_path)[0] + '_meta.json'
    json.dump(meta_out, open(meta_path, 'w', encoding='utf-8'), ensure_ascii=False, indent=2, default=str)

    print(f'\n完成！已輸出：{out_path}')
    print(f'結構資訊已輸出：{meta_path}（給估驗計價/變更設計匯入工具用）')


if __name__ == '__main__':
    main()
