# -*- coding: utf-8 -*-
"""
施工日報版面模組（依「附表四_公共工程施工日誌」官方逐格核對重建，見專案筆記成果二說明）。

這個檔案不是獨立可執行的腳本，是被 build_construction_log.py 用 exec() 在同一個
命名空間下執行的「版面樣板」，執行前呼叫端要先把下列變數放進命名空間：
  wb, FIRST_ITEM_COL, N_ITEMS, FIRST_MAT_COL, N_MAT, FIRST_LAB_COL, N_LABOR,
  FIRST_MCH_COL, N_MACH, ROW_SEQ, ROW_NAME, ROW_UNIT, ROW_QTY, ROW_PRICE, ROW_LABEL,
  ROW_REMARK, FIRST_DAY_ROW, LAST_DAY_ROW, FIRST_TXT_COL, TXT_FIELDS, CATEGORIES

完全通用：本檔案本身不含任何工程專屬的硬編碼數字，所有版面尺寸都是依 N_ITEMS/
CATEGORIES/N_MAT/N_LABOR/N_MACH 動態算出，跟原始 build_v4_report.py（原本讀
_v4_meta.json）邏輯完全一致，只是改成讀記憶體中的命名空間變數。
"""
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import Outline

item_cols = [get_column_letter(FIRST_ITEM_COL + i) for i in range(N_ITEMS)]
mat_cols = [get_column_letter(FIRST_MAT_COL + i) for i in range(N_MAT)]
lab_cols = [get_column_letter(FIRST_LAB_COL + i) for i in range(N_LABOR)]
mch_cols = [get_column_letter(FIRST_MCH_COL + i) for i in range(N_MACH)]
field_map = {key: get_column_letter(FIRST_TXT_COL + i) for i, (key, _, _) in enumerate(TXT_FIELDS)}

ws = wb.create_sheet('施工日報')

MED = 'medium'; THN = 'thin'
def _side(style): return Side(style=style, color='000000')

def box(r1, c1, r2, c2):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            top = MED if r == r1 else THN
            bottom = MED if r == r2 else THN
            left = MED if c == c1 else THN
            right = MED if c == c2 else THN
            cell.border = Border(top=_side(top), bottom=_side(bottom), left=_side(left), right=_side(right))

def box_thick(r1, c1, r2, c2):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = Border(
                top=_side('thick'), bottom=_side('thick'), left=_side('thick'), right=_side('thick'))

FONT_NAME = '新細明體'
bold = Font(bold=True, name=FONT_NAME, size=12)
normal = Font(name=FONT_NAME, size=12)
small = Font(name=FONT_NAME, size=10)
title_font = Font(name=FONT_NAME, size=16)
note_font = Font(italic=True, color='C00000', size=9, name=FONT_NAME)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_mid = Alignment(horizontal='left', vertical='center', wrap_text=True)
left_top = Alignment(horizontal='left', vertical='top', wrap_text=True)
left_mid_nowrap = Alignment(horizontal='left', vertical='center', wrap_text=False)

TOTAL_COLS = 12
LEFT_START_COL = 1
RIGHT_START_COL = 7
BLOCK_WIDTH = 6

DATE_CELL_COL = 'G'
DATE_CELL_ROW = 4
DATE_CELL = f'{DATE_CELL_COL}{DATE_CELL_ROW}'

def matchidx():
    return f"MATCH(${DATE_CELL_COL}${DATE_CELL_ROW},'日資料庫'!$B${FIRST_DAY_ROW}:$B${LAST_DAY_ROW},0)"

def lookup(col_letter):
    return f"INDEX('日資料庫'!${col_letter}${FIRST_DAY_ROW}:${col_letter}${LAST_DAY_ROW},{matchidx()})"

def lookup_blankaware(col_letter):
    idx = lookup(col_letter)
    return f'=IFERROR(IF({idx}=0,"",{idx}),"")'

def lookup_blankaware_expr(col_letter):
    idx = lookup(col_letter)
    return f'IFERROR(IF({idx}=0,"",{idx}),"")'

def cum(col_letter):
    return (f"=IFERROR(SUM('日資料庫'!${col_letter}${FIRST_DAY_ROW}:"
            f"INDEX('日資料庫'!${col_letter}${FIRST_DAY_ROW}:${col_letter}${LAST_DAY_ROW},{matchidx()})),\"\")")

# ============ Row 1：標題 ============
ws.merge_cells(f'A1:{get_column_letter(TOTAL_COLS)}1')
ws['A1'] = '公共工程施工日誌'
ws['A1'].font = title_font
ws['A1'].alignment = Alignment(horizontal='center')

# ============ Row 2：非表格提示列 ============
ws.merge_cells(f'A2:{get_column_letter(TOTAL_COLS)}2')
ws['A2'] = f'※{DATE_CELL}「填表日期」請改成要查詢/列印的日期；開工日期、工期、材料/工別/機具名稱僅需在「日資料庫」工作表設定一次，此處自動帶入。'
ws['A2'].font = note_font

# ============ Row 3：表報編號 ============
ws.row_dimensions[3].height = 18
ws['A3'] = '表報編號：'; ws['A3'].font = bold; ws['A3'].alignment = left_mid_nowrap
ws.merge_cells('B3:D3')
ws['B3'].fill = PatternFill('solid', fgColor='FFF2CC')
ws['B3'].font = normal; ws['B3'].alignment = left_mid_nowrap

# ============ Row 4：本日天氣／填表日期（星期） ============
ws.row_dimensions[DATE_CELL_ROW].height = 18
weather_prefix = (f'="本日天氣：上午："&{lookup_blankaware_expr("D")}&"　下午："&'
                   f'{lookup_blankaware_expr("E")}&"　　　　填表日期："')
ws.merge_cells(f'A{DATE_CELL_ROW}:F{DATE_CELL_ROW}')
wcell = ws.cell(row=DATE_CELL_ROW, column=1, value=weather_prefix)
wcell.font = normal; wcell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=False)

ws.merge_cells(f'G{DATE_CELL_ROW}:H{DATE_CELL_ROW}')
dcell = ws[DATE_CELL]
dcell.value = "='日資料庫'!B2"
dcell.number_format = 'yyyy"年"m"月"d"日"'
dcell.fill = PatternFill('solid', fgColor='FFF2CC')
dcell.font = normal; dcell.alignment = left_mid_nowrap

ws.merge_cells(f'I{DATE_CELL_ROW}:{get_column_letter(TOTAL_COLS)}{DATE_CELL_ROW}')
wk = ws.cell(row=DATE_CELL_ROW, column=9,
             value=f'="（星期"&CHOOSE(WEEKDAY({DATE_CELL},2),"一","二","三","四","五","六","日")&"）"')
wk.font = normal; wk.alignment = left_mid_nowrap

# ============ 表頭資訊框：逐列比照官方附表四原稿(4列，medium外框/thin內格) ============
_col = lambda c: get_column_letter(c)
R_HDR_A = DATE_CELL_ROW + 1
R_HDR_B = R_HDR_A + 1
R_HDR_C = R_HDR_B + 1
R_HDR_D = R_HDR_C + 1

def place_field(row, lc1, lc2, vc1, vc2, label, value, fmt=None):
    if lc2 > lc1:
        ws.merge_cells(start_row=row, start_column=lc1, end_row=row, end_column=lc2)
    lcell = ws.cell(row=row, column=lc1, value=label)
    lcell.font = bold; lcell.alignment = left_mid_nowrap
    if vc2 > vc1:
        ws.merge_cells(start_row=row, start_column=vc1, end_row=row, end_column=vc2)
    vcell = ws.cell(row=row, column=vc1, value=value)
    vcell.font = normal; vcell.alignment = left_mid_nowrap
    if fmt:
        vcell.number_format = fmt
    return vcell

place_field(R_HDR_A, 1, 2, 3, 6, '工程名稱', "='日資料庫'!B1")
place_field(R_HDR_A, 7, 8, 9, 12, '承攬廠商名稱', '')

place_field(R_HDR_B, 1, 1, 2, 3, '核定工期', "='日資料庫'!D2", '0"天"')
place_field(R_HDR_B, 4, 4, 5, 6, '累計工期', f'={DATE_CELL}-{_col(3)}{R_HDR_C}+1', '0"天"')
place_field(R_HDR_B, 7, 7, 8, 9, '剩餘工期', f'={_col(2)}{R_HDR_B}-{_col(5)}{R_HDR_B}', '0"天"')
place_field(R_HDR_B, 10, 11, 12, 12, '工期展延天數', "='日資料庫'!M1", '0"天"')

place_field(R_HDR_C, 1, 2, 3, 6, '開工日期', "='日資料庫'!B2", 'yyyy"年"m"月"d"日"')
place_field(R_HDR_C, 7, 8, 9, 12, '完工日期', "='日資料庫'!F2", 'yyyy"年"m"月"d"日"')

sched_cell = place_field(R_HDR_D, 1, 2, 3, 4, '預定進度(%)', lookup_blankaware(field_map['SCHED']), '0.00%')
actual_cell = place_field(R_HDR_D, 5, 6, 7, 8, '實際進度(%)', f"={lookup(CATEGORIES[0][4])}", '0.00%')
# 超前/落後 = 實際進度 - 預定進度，直接參照上面兩個已放好的儲存格，避免重複計算INDEX/MATCH。
# 兩者皆為空字串（尚無資料）時同樣顯示空白；只要有一邊有值就照算，正值代表超前、負值代表落後。
_diff_formula = (f'=IFERROR(IF(OR({sched_cell.coordinate}="",{actual_cell.coordinate}=""),"",'
                  f'{actual_cell.coordinate}-{sched_cell.coordinate}),"")')
place_field(R_HDR_D, 9, 10, 11, 12, '超前/落後(%)', _diff_formula, '+0.00%;-0.00%;0.00%')

box(R_HDR_A, 1, R_HDR_D, TOTAL_COLS)
for r in (R_HDR_A, R_HDR_B, R_HDR_C, R_HDR_D):
    ws.row_dimensions[r].height = 20

R_AFTER_HDR = R_HDR_D

# ============ 補充：工程編號／主辦機關／契約總價／施工地點 ============
R_EXTRA = R_AFTER_HDR + 1
ws.merge_cells(start_row=R_EXTRA, start_column=1, end_row=R_EXTRA, end_column=TOTAL_COLS)
extra_txt = ('="（本系統擴充，非附表四固定欄位）工程編號："&\'日資料庫\'!G1&"　　主辦機關："&\'日資料庫\'!K1&'
             '"　　契約總價："&TEXT(\'日資料庫\'!I1,"#,##0")&"　　施工地點："&\'日資料庫\'!H2')
ec = ws.cell(row=R_EXTRA, column=1, value=extra_txt)
ec.font = small; ec.alignment = left_mid_nowrap
ws.row_dimensions[R_EXTRA].height = 16

R_AFTER_HDR = R_EXTRA

# ============ 施工進度統計 ============
# 中項筆數常常較多、逐項列出會佔用大量畫面/列印空間，這裡把「總計＋大項」跟「中項明細」
# 拆成兩個區塊：大項彙總一律顯示，中項明細另外用Excel列大綱分組、預設摺疊收起（資料仍在，
# 只是先收合，需要細看時自行展開），呼應使用者反映「中項有點佔畫面空間」的需求。
prog_cols_block = ['類別', '名稱', '契約金額', '累計完成率']

def _cat_level(label_text):
    ch = label_text[0] if label_text else ''
    if ch in TOP_NUMERALS:
        return 1
    if ch in MID_NUMERALS:
        return 2
    return 0  # 總計，或偵測不到編號的類別

def _prog_tag_name(label_text):
    if '、' in label_text:
        tag, name = label_text.split('、', 1)
    elif label_text.startswith('總計'):
        tag, name = '總計', '全案'
    else:
        tag, name = label_text, ''
    return tag, (name if name else label_text)

def _write_prog_block(cats, head_row, bold_all):
    for base in (LEFT_START_COL, RIGHT_START_COL):
        for j, cname in enumerate(prog_cols_block):
            c = ws.cell(row=head_row, column=base + j, value=cname)
            c.font = bold; c.alignment = center
    half = (len(cats) + 1) // 2
    data_start = head_row + 1
    for i in range(half):
        rr = data_start + i
        for base, idx in ((LEFT_START_COL, i), (RIGHT_START_COL, i + half)):
            if idx >= len(cats):
                continue
            label_text, start_no, end_no, total_price, col_letter = cats[idx]
            tag, name = _prog_tag_name(label_text)
            vals = [tag, name, total_price, f"={lookup(col_letter)}"]
            for j, val in enumerate(vals):
                c = ws.cell(row=rr, column=base + j, value=val)
                c.font = bold if (bold_all or idx == 0) else small
                c.alignment = left_mid if j == 1 else center
                if j == 2: c.number_format = '#,##0'
                if j == 3: c.number_format = '0.00%'
    end_row = data_start + half - 1
    box(head_row, 1, end_row, TOTAL_COLS)
    box(head_row, 1, end_row, BLOCK_WIDTH)
    box(head_row, RIGHT_START_COL, end_row, TOTAL_COLS)
    return end_row

top_cats = [c for c in CATEGORIES if _cat_level(c[0]) in (0, 1)]
mid_cats = [c for c in CATEGORIES if _cat_level(c[0]) == 2]

PROG_TITLE_ROW = R_AFTER_HDR + 1
ws.merge_cells(start_row=PROG_TITLE_ROW, start_column=1, end_row=PROG_TITLE_ROW, end_column=TOTAL_COLS)
ws.cell(row=PROG_TITLE_ROW, column=1, value='施工進度統計（按契約金額加權累計完成率，非官方日誌固定欄位，屬本系統擴充分析）：').font = bold

TOP_HEAD_ROW = PROG_TITLE_ROW + 1
TOP_END_ROW = _write_prog_block(top_cats, TOP_HEAD_ROW, bold_all=True)
PROG_END_ROW = TOP_END_ROW

if mid_cats:
    MID_TOGGLE_ROW = TOP_END_ROW + 1
    ws.merge_cells(start_row=MID_TOGGLE_ROW, start_column=1, end_row=MID_TOGGLE_ROW, end_column=TOTAL_COLS)
    tgl = ws.cell(row=MID_TOGGLE_ROW, column=1,
                  value=f'中項明細（共{len(mid_cats)}項，預設已摺疊以節省畫面/列印空間；'
                        f'點選下方列號左側大綱區的「+」展開、「－」摺疊，資料都還在）：')
    tgl.font = note_font
    box(MID_TOGGLE_ROW, 1, MID_TOGGLE_ROW, TOTAL_COLS)
    MID_HEAD_ROW = MID_TOGGLE_ROW + 1
    MID_END_ROW = _write_prog_block(mid_cats, MID_HEAD_ROW, bold_all=False)
    PROG_END_ROW = MID_END_ROW

# 使用者反映：這整段「（本系統擴充）工程編號/主辦機關/契約總價/施工地點」補充列＋
# 「施工進度統計」彙總表（含上面才展開的中項明細子表），並非附表四官方固定欄位，
# 畫面/列印時預設整段摺疊隱藏即可，需要查看時再從大綱區展開──此範圍是動態算出來的
# (R_EXTRA ~ PROG_END_ROW)，不是寫死的列號，這樣不論案子有幾個大項/中項，永遠都是
# 「施工進度統計整段」被摺疊，行為一致。
ws.sheet_properties.outlinePr = Outline(summaryBelow=False)
for r in range(R_EXTRA, PROG_END_ROW + 1):
    ws.row_dimensions[r].outlineLevel = 1
    ws.row_dimensions[r].hidden = True

# ============ 一、施工項目表 ============
SEC1_ROW = PROG_END_ROW + 1
ws.merge_cells(start_row=SEC1_ROW, start_column=1, end_row=SEC1_ROW, end_column=TOTAL_COLS)
ws.cell(row=SEC1_ROW, column=1, value='一、依施工計畫書執行按圖施工概況（含約定之重要施工項目及完成數量等）：').font = bold

TABLE_HEAD_ROW = SEC1_ROW + 1
# 使用者反映不需要「備註」欄，已拿掉；拿掉後空出來的最後一欄併給「累計完成數量」，
# 讓每個區塊仍維持原本BLOCK_WIDTH(6)欄寬，跟上下其他表格(材料/人員機具等)的12欄
# 版面對齊，不會留白也不用改動BLOCK_WIDTH這個其他區塊也共用的全域常數。
cols_block = ['施工項目', '單位', '契約數量', '本日完成數量', '累計完成數量']
LAST_COL_IDX = len(cols_block) - 1
for base in (LEFT_START_COL, RIGHT_START_COL):
    for j, cname in enumerate(cols_block):
        col = base + j
        if j == LAST_COL_IDX:
            ws.merge_cells(start_row=TABLE_HEAD_ROW, start_column=col,
                            end_row=TABLE_HEAD_ROW, end_column=base + BLOCK_WIDTH - 1)
        c = ws.cell(row=TABLE_HEAD_ROW, column=col, value=cname)
        c.font = bold; c.alignment = center

HALF = (N_ITEMS + 1) // 2
DATA_START_ROW = TABLE_HEAD_ROW + 1
for i in range(HALF):
    rr = DATA_START_ROW + i
    for base, item_idx in ((LEFT_START_COL, i), (RIGHT_START_COL, i + HALF)):
        if item_idx >= N_ITEMS:
            continue
        cl = item_cols[item_idx]
        vals_formula = [
            f"='日資料庫'!{cl}{ROW_NAME}",
            f"='日資料庫'!{cl}{ROW_UNIT}",
            f"='日資料庫'!{cl}{ROW_QTY}",
            f"=IFERROR({lookup(cl)},\"\")",
            cum(cl),
        ]
        for j, val in enumerate(vals_formula):
            col = base + j
            if j == LAST_COL_IDX:
                ws.merge_cells(start_row=rr, start_column=col, end_row=rr, end_column=base + BLOCK_WIDTH - 1)
            c = ws.cell(row=rr, column=col, value=val)
            c.font = small
            c.alignment = left_mid if j == 0 else center
            if j in (2, 3, 4):
                c.number_format = '#,##0.##'
TABLE_END_ROW = DATA_START_ROW + HALF - 1
box(SEC1_ROW, 1, TABLE_END_ROW, TOTAL_COLS)
box(TABLE_HEAD_ROW, 1, TABLE_END_ROW, BLOCK_WIDTH)
box(TABLE_HEAD_ROW, RIGHT_START_COL, TABLE_END_ROW, TOTAL_COLS)

cur_row = TABLE_END_ROW + 1

# ============ 「營造業專業工程特定施工項目 A./B.」 ============
AB_TITLE_ROW = cur_row
ws.merge_cells(start_row=AB_TITLE_ROW, start_column=1, end_row=AB_TITLE_ROW, end_column=TOTAL_COLS)
ws.cell(row=AB_TITLE_ROW, column=1, value='營造業專業工程特定施工項目').font = bold
for key, label in (('SEC1_A', 'A.'), ('SEC1_B', 'B.')):
    r = AB_TITLE_ROW + (1 if key == 'SEC1_A' else 2)
    ws.cell(row=r, column=1, value=label).font = normal
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=TOTAL_COLS)
    c = ws.cell(row=r, column=2, value=lookup_blankaware(field_map[key]))
    c.font = small
AB_END_ROW = AB_TITLE_ROW + 2
box_thick(AB_TITLE_ROW, 1, AB_END_ROW, TOTAL_COLS)
cur_row = AB_END_ROW + 1

# ============ 二、工地材料管理概況 ============
SEC2_ROW = cur_row
ws.merge_cells(start_row=SEC2_ROW, start_column=1, end_row=SEC2_ROW, end_column=TOTAL_COLS)
ws.cell(row=SEC2_ROW, column=1, value='二、工地材料管理概況（含約定之重要材料使用狀況及數量等）：').font = bold

MAT_HEAD_ROW = SEC2_ROW + 1
mat_cols_block = ['材料名稱', '單位', '契約數量', '本日使用數量', '累計使用數量', '備註']
for base in (LEFT_START_COL, RIGHT_START_COL):
    for j, cname in enumerate(mat_cols_block):
        c = ws.cell(row=MAT_HEAD_ROW, column=base + j, value=cname)
        c.font = bold; c.alignment = center

MAT_HALF = (N_MAT + 1) // 2
MAT_DATA_START = MAT_HEAD_ROW + 1
for i in range(MAT_HALF):
    rr = MAT_DATA_START + i
    for base, m_idx in ((LEFT_START_COL, i), (RIGHT_START_COL, i + MAT_HALF)):
        if m_idx >= N_MAT:
            continue
        cl = mat_cols[m_idx]
        name_ref = f"'日資料庫'!{cl}{ROW_NAME}"
        vals = [
            f'=IF({name_ref}="","",{name_ref})',
            f'=IF({name_ref}="","",\'日資料庫\'!{cl}{ROW_UNIT})',
            f'=IF({name_ref}="","",\'日資料庫\'!{cl}{ROW_QTY})',
            f"=IFERROR(IF('日資料庫'!{cl}{ROW_NAME}=\"\",\"\",{lookup(cl)}),\"\")",
            f"=IFERROR(IF('日資料庫'!{cl}{ROW_NAME}=\"\",\"\",{cum(cl)[1:]}),\"\")",
            f'=IF(\'日資料庫\'!{cl}{ROW_REMARK}="","",\'日資料庫\'!{cl}{ROW_REMARK})',
        ]
        for j, val in enumerate(vals):
            cc = ws.cell(row=rr, column=base + j, value=val)
            cc.font = small
            cc.alignment = left_mid if j in (0, 5) else center
            if j in (2, 3, 4): cc.number_format = '#,##0.##'
MAT_END_ROW = MAT_DATA_START + MAT_HALF - 1
box(SEC2_ROW, 1, MAT_END_ROW, TOTAL_COLS)
box(MAT_HEAD_ROW, 1, MAT_END_ROW, BLOCK_WIDTH)
box(MAT_HEAD_ROW, RIGHT_START_COL, MAT_END_ROW, TOTAL_COLS)

cur_row = MAT_END_ROW + 1

# ============ 三、工地人員及機具管理 ============
SEC3_ROW = cur_row
ws.merge_cells(start_row=SEC3_ROW, start_column=1, end_row=SEC3_ROW, end_column=TOTAL_COLS)
ws.cell(row=SEC3_ROW, column=1, value='三、工地人員及機具管理（含約定之出工人數及機具使用情形及數量）：').font = bold

LM_HEAD_ROW = SEC3_ROW + 1
labels_labor = ['工別', '本日人數', '累計人數']
labels_mach = ['機具名稱', '本日使用數量', '累計使用數量']
for j, cname in enumerate(labels_labor):
    c = ws.cell(row=LM_HEAD_ROW, column=LEFT_START_COL + j, value=cname); c.font = bold; c.alignment = center
for j, cname in enumerate(labels_labor):
    c = ws.cell(row=LM_HEAD_ROW, column=LEFT_START_COL + 3 + j, value=cname); c.font = bold; c.alignment = center
for j, cname in enumerate(labels_mach):
    c = ws.cell(row=LM_HEAD_ROW, column=RIGHT_START_COL + j, value=cname); c.font = bold; c.alignment = center
for j, cname in enumerate(labels_mach):
    c = ws.cell(row=LM_HEAD_ROW, column=RIGHT_START_COL + 3 + j, value=cname); c.font = bold; c.alignment = center

LM_ROWS = max(N_LABOR, N_MACH)
LM_DATA_START = LM_HEAD_ROW + 1
LAB_HALF = (N_LABOR + 1) // 2
MCH_HALF = (N_MACH + 1) // 2
for i in range(max(LAB_HALF, MCH_HALF)):
    rr = LM_DATA_START + i
    for grp_base, l_idx in ((LEFT_START_COL, i), (LEFT_START_COL + 3, i + LAB_HALF)):
        if l_idx >= N_LABOR:
            continue
        cl = lab_cols[l_idx]
        name_ref = f"'日資料庫'!{cl}{ROW_NAME}"
        vals = [f'=IF({name_ref}="","",{name_ref})',
                f"=IFERROR(IF('日資料庫'!{cl}{ROW_NAME}=\"\",\"\",{lookup(cl)}),\"\")",
                f"=IFERROR(IF('日資料庫'!{cl}{ROW_NAME}=\"\",\"\",{cum(cl)[1:]}),\"\")"]
        for j, val in enumerate(vals):
            c = ws.cell(row=rr, column=grp_base + j, value=val)
            c.font = small; c.alignment = left_mid if j == 0 else center
    for grp_base, m_idx in ((RIGHT_START_COL, i), (RIGHT_START_COL + 3, i + MCH_HALF)):
        if m_idx >= N_MACH:
            continue
        cl = mch_cols[m_idx]
        name_ref = f"'日資料庫'!{cl}{ROW_NAME}"
        vals = [f'=IF({name_ref}="","",{name_ref})',
                f"=IFERROR(IF('日資料庫'!{cl}{ROW_NAME}=\"\",\"\",{lookup(cl)}),\"\")",
                f"=IFERROR(IF('日資料庫'!{cl}{ROW_NAME}=\"\",\"\",{cum(cl)[1:]}),\"\")"]
        for j, val in enumerate(vals):
            c = ws.cell(row=rr, column=grp_base + j, value=val)
            c.font = small; c.alignment = left_mid if j == 0 else center
LM_END_ROW = LM_DATA_START + max(LAB_HALF, MCH_HALF) - 1
box(SEC3_ROW, 1, LM_END_ROW, TOTAL_COLS)
box(LM_HEAD_ROW, 1, LM_END_ROW, BLOCK_WIDTH)
box(LM_HEAD_ROW, RIGHT_START_COL, LM_END_ROW, TOTAL_COLS)

cur_row = LM_END_ROW + 1

# ============ 四~八：自由文字/勾選欄位 ============
def text_section(cur_row, title, col_letter, value_rows=2):
    ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=TOTAL_COLS)
    ws.cell(row=cur_row, column=1, value=title).font = bold
    body_row = cur_row + 1
    ws.merge_cells(start_row=body_row, start_column=1, end_row=body_row + value_rows - 1, end_column=TOTAL_COLS)
    c = ws.cell(row=body_row, column=1, value=lookup_blankaware(col_letter))
    c.font = small; c.alignment = left_top
    box(cur_row, 1, body_row + value_rows - 1, TOTAL_COLS)
    return body_row + value_rows

def yn_checkbox_expr(col_letter, options):
    # 依「日資料庫」該欄目前的值(有/無/無新進勞工…)，動態組出「☑有 □無」樣式的純文字，
    # 用□(未勾)／☑(已勾)兩種Unicode符號模擬勾選框；找不到對應選項(含空白/其他文字)時，
    # 一律回退為全部未勾選的樣子，不猜測、不強行對應到某個選項。
    raw = lookup(col_letter)
    def row_str(selected):
        return '  '.join(f'☑{opt}' if opt == selected else f'□{opt}' for opt in options)
    default_str = row_str(None)
    expr = f'"{default_str}"'
    for opt in reversed(options):
        expr = f'IF({raw}="{opt}","{row_str(opt)}",{expr})'
    return f'=IFERROR({expr},"{default_str}")', default_str

# 勾選框樣式的顯示區改用較窄字級(small)以容納較長文字(尤其是含「無新進勞工」的3選項版本)，
# 並把數值欄從2欄加寬為4欄，標題欄相應從10欄縮為8欄。
YN_TITLE_END = TOTAL_COLS - 4   # 8
YN_VALUE_START = TOTAL_COLS - 3  # 9

def yn_section(cur_row, title, col_letter, options=('有', '無')):
    ws.cell(row=cur_row, column=1, value=title).font = normal
    ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=YN_TITLE_END)
    expr, _default_str = yn_checkbox_expr(col_letter, options)
    c = ws.cell(row=cur_row, column=YN_VALUE_START, value=expr)
    c.font = small; c.alignment = center
    ws.merge_cells(start_row=cur_row, start_column=YN_VALUE_START, end_row=cur_row, end_column=TOTAL_COLS)
    box(cur_row, 1, cur_row, TOTAL_COLS)
    return cur_row + 1

SEC4_ROW = cur_row
ws.merge_cells(start_row=SEC4_ROW, start_column=1, end_row=SEC4_ROW + 1, end_column=TOTAL_COLS - 1)
ws.cell(row=SEC4_ROW, column=1,
        value='四、本日施工項目是否有須依「營造業專業工程特定施工項目應置之技術士種類、比率或人數標準表」'
              '規定應設置技術士之專業工程：（此項如勾選"有"，則應填寫後附「公共工程施工日誌之技術士簽章表」）')
c4 = ws.cell(row=SEC4_ROW, column=1)
c4.font = normal; c4.alignment = left_top
ws.merge_cells(start_row=SEC4_ROW, start_column=TOTAL_COLS, end_row=SEC4_ROW + 1, end_column=TOTAL_COLS)
v4 = ws.cell(row=SEC4_ROW, column=TOTAL_COLS, value=lookup_blankaware(field_map['SEC4']))
v4.font = bold; v4.alignment = center
box_thick(SEC4_ROW, 1, SEC4_ROW + 1, TOTAL_COLS)
cur_row = SEC4_ROW + 2

SEC5_TITLE_ROW = cur_row
ws.merge_cells(start_row=SEC5_TITLE_ROW, start_column=1, end_row=SEC5_TITLE_ROW, end_column=TOTAL_COLS)
ws.cell(row=SEC5_TITLE_ROW, column=1, value='五、工地職業安全衛生事項之督導、公共環境與安全之維護及其他工地行政事務：').font = bold
r = SEC5_TITLE_ROW + 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=TOTAL_COLS)
ws.cell(row=r, column=1, value='(一)施工前檢查事項：').font = bold
r += 1
r = yn_section(r, '  1.實施勤前教育(含工地預防災變及危害告知)：', field_map['SEC5A'])
# SEC5B在「日資料庫」的資料驗證清單是"有,無,無新進勞工"(dv_yn2)，這裡的勾選框選項順序與其一致。
r = yn_section(r, '  2.確認新進勞工是否提報勞工保險(或其他商業保險)資料及安全衛生教育訓練紀錄：', field_map['SEC5B'],
               options=('有', '無', '無新進勞工'))
r = yn_section(r, '  3.檢查勞工個人防護具：', field_map['SEC5C'])
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=TOTAL_COLS)
ws.cell(row=r, column=1, value='(二)其他事項：').font = bold
r += 1
body_row = r
ws.merge_cells(start_row=body_row, start_column=1, end_row=body_row + 1, end_column=TOTAL_COLS)
c = ws.cell(row=body_row, column=1, value=lookup_blankaware(field_map['SEC5_OTHER']))
c.font = small; c.alignment = left_top
box(SEC5_TITLE_ROW, 1, body_row + 1, TOTAL_COLS)
cur_row = body_row + 2

cur_row = text_section(cur_row, '六、施工取樣試驗紀錄：', field_map['SEC6'], value_rows=2)
cur_row = text_section(cur_row, '七、通知協力廠商辦理事項：', field_map['SEC7'], value_rows=2)
cur_row = text_section(cur_row, '八、重要事項記錄：', field_map['SEC8'], value_rows=3)

# ============ 簽章欄 ============
SIGN_ROW = cur_row
ws.cell(row=SIGN_ROW, column=1, value='簽章：【工地主任】（註3）：').font = bold
SIGN_END_ROW = SIGN_ROW + 6
box(SIGN_ROW, 1, SIGN_END_ROW, TOTAL_COLS)
ws.merge_cells(start_row=SIGN_ROW, start_column=1, end_row=SIGN_ROW, end_column=TOTAL_COLS)

FINAL_ROW = SIGN_END_ROW

# ============ 版面設定：直向A4、無底色、無格線，寬1頁 x 高最多2頁 ============
# 欄寬依「紙張可用寬度」等比例縮放，讓畫面上看到的欄寬比例跟印出來的樣子一致，而不是單靠
# fitToWidth=1硬壓縮：
#   假設1：欄寬單位換算比照Excel預設(工作表未特別設定「常用」樣式字型時)的經驗值，
#          約1欄寬單位≈7px＠96DPI(此換算是Excel官方文件給的概略值，實際會因作業系統
#          字型引擎微幅不同，不是每台電腦都逐像素一致)。
#   假設2：紙張為A4直向(8.27吋寬)，扣掉下面左右邊界(左0.4吋+右0.2吋=0.6吋)後，
#          可用寬度＝8.27-0.6=7.67吋＝7.67×96÷7≈105欄寬單位。
#   做法：把原本的欄寬(總和149單位)等比例縮到105單位(比例≈0.705)，四捨五入取整數，
#          縮完後總和剛好105，跟可用寬度一致；相對比例維持不變，不會有的欄變得不成比例窄。
# 邊界維持既有設定(左0.4/右0.2/上0.4/下0.4吋，頁首頁尾0)，因使用者這次沒有給出具體邊界數字，
# 就不擅自改回Excel出廠的預設值(0.7/0.7/0.75/0.75吋)，以免跟原本已核可的版面衝突；
# fitToWidth=1仍保留當作最後一道保險，萬一實際字型換算跟假設有落差，列印仍保證縮放到剛好一頁寬。
widths = {1: 11, 2: 9, 3: 11, 4: 9, 5: 9, 6: 7, 7: 11, 8: 7, 9: 9, 10: 8, 11: 7, 12: 7}
for col, w in widths.items():
    ws.column_dimensions[get_column_letter(col)].width = w

ws.sheet_view.showGridLines = False
ws.print_area = f'A1:{get_column_letter(TOTAL_COLS)}{FINAL_ROW}'
ws.page_setup.orientation = 'portrait'
ws.page_setup.paperSize = ws.PAPERSIZE_A4
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.page_setup.fitToPage = True
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 2
ws.page_margins.left = 0.4
ws.page_margins.right = 0.2
ws.page_margins.top = 0.4
ws.page_margins.bottom = 0.4
ws.page_margins.header = 0
ws.page_margins.footer = 0

# 頂端列印標題：第1~TABLE_HEAD_ROW列(官方4列表頭框＋本系統擴充的施工進度統計整段(已摺疊隱藏，
# 隱藏部分本來就不會印出)＋「一、施工項目表」標題列＋該表格自己的欄名列)，每頁都重複列印，
# 這樣當「一、施工項目表」內容超過一頁時，第2頁開頭仍會看到「施工項目/單位/契約數量/…」欄名，
# 不會變成一片沒有欄名的數字。這個範圍同樣是用TABLE_HEAD_ROW動態算出來，不是寫死23。
ws.print_title_rows = f'1:{TABLE_HEAD_ROW}'
